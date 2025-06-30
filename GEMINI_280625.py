# <<< YENİ EKLEME: Excel işlemleri için >>>
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
# <<< SON >>>

from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QComboBox, QLineEdit, QListWidget, QMessageBox, QCheckBox, QGridLayout,
    QInputDialog, QFormLayout, QScrollArea, QFileDialog, QTextEdit, QProgressDialog
)
from PyQt5.QtCore import Qt
import json
import os
import random # <<< YENİ EKLEME: Ders yerleştirmede rastgelelik için >>>
import math # <<< YENİ EKLEME: Excel sayfa ayarı için >>>


class SchoolSchedulerApp(QWidget):
    def __init__(self):
        super().__init__()
        self.school_level = None
        self.classes = {}
        self.selected_class = None
        self.daily_hours = 0
        self.last_run_unplaced_lessons = []
        self.courses = [
            "Yabancı Dil (İngilizce)", "İkinci Yabancı Dil (Almanca)", "Türkçe", "Tarih",
             "Fizik", "Kimya", "Biyoloji", "Matematik", "Felsefe", "Görsel Sanatlar",
             "Beden Eğitimi", "Coğrafya", 
             "Din Kültürü ve Ahlak Bilgisi", "Seçmeli İngilizce"
        ]
        self.teachers = {}
        self.current_teacher = None
        self.current_class = None

        # <<< YENİ EKLEME: Haftanın günleri (program oluşturma için) >>>
        self.days_of_week = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]

        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout() # Ana layout

        # Üst kısım (Veri Girişi ve Yönetimi) için bir widget ve layout
        top_widget = QWidget()
        layout = QVBoxLayout(top_widget)

        # --- Mevcut UI (layout üzerine eklenecek) ---
        file_buttons_layout = QHBoxLayout()
        self.show_template_btn = QPushButton("Formu Göster")
        self.show_template_btn.clicked.connect(self.generate_template)
        self.load_template_btn = QPushButton("Notepad'den Al")
        self.load_template_btn.clicked.connect(self.load_from_file)
        file_buttons_layout.addWidget(self.show_template_btn)
        file_buttons_layout.addWidget(self.load_template_btn)
        layout.addLayout(file_buttons_layout)

        save_load_layout = QHBoxLayout()
        self.save_data_btn = QPushButton("Tüm Bilgileri Kaydet")
        self.save_data_btn.clicked.connect(self.save_all_data)
        self.load_data_btn = QPushButton("Kayıtlı Verileri Yükle")
        self.load_data_btn.clicked.connect(self.load_saved_data)
        save_load_layout.addWidget(self.save_data_btn)
        save_load_layout.addWidget(self.load_data_btn)
        layout.addLayout(save_load_layout)

        # ... (Diğer UI elemanları layout'a eklenmeye devam eder) ...
        self.label = QLabel("Okul Kademesi Seçin:")
        layout.addWidget(self.label)
        self.school_level_dropdown = QComboBox()
        self.school_level_dropdown.addItems(["Lise", "Ortaokul", "İlkokul"])
        self.school_level_dropdown.currentTextChanged.connect(self.set_school_level)
        layout.addWidget(self.school_level_dropdown)

        self.section_inputs = {}
        self.section_layout = QGridLayout()
        layout.addLayout(self.section_layout)

        self.daily_hours_label = QLabel("Günlük Ders Saati:")
        self.daily_hours_input = QLineEdit()
        self.daily_hours_input.setPlaceholderText("Günlük ders saatinizi girin (Sayısal)")
        layout.addWidget(self.daily_hours_label)
        layout.addWidget(self.daily_hours_input)

        self.create_classes_button = QPushButton("Şubeleri Oluştur")
        self.create_classes_button.clicked.connect(self.create_classes)
        layout.addWidget(self.create_classes_button)

        # Orta Bölüm (Listeler) için QHBoxLayout
        list_layout = QHBoxLayout()

        # Sınıf Listesi ve Butonları (Sol Taraf)
        class_area_layout = QVBoxLayout()
        class_area_layout.addWidget(QLabel("Sınıflar:"))
        self.class_list = QListWidget()
        self.class_list.itemClicked.connect(self.select_class)
        class_area_layout.addWidget(self.class_list)
        # Sınıf Yönetim Butonları
        class_buttons_layout = QGridLayout() # Butonları grid ile dizelim
        self.assign_courses_button = QPushButton("Dersleri Ata")
        self.assign_courses_button.clicked.connect(self.assign_courses)
        class_buttons_layout.addWidget(self.assign_courses_button, 0, 0)
        self.set_hours_button = QPushButton("Saatleri Ayarla")
        self.set_hours_button.clicked.connect(self.set_course_hours)
        class_buttons_layout.addWidget(self.set_hours_button, 0, 1)
        self.remove_course_button = QPushButton("Sınıftan Ders Sil")
        self.remove_course_button.clicked.connect(self.remove_course)
        class_buttons_layout.addWidget(self.remove_course_button, 1, 0)
        self.add_course_button = QPushButton("Genel Ders Ekle") # İsmi kısalttık
        self.add_course_button.clicked.connect(self.add_course)
        class_buttons_layout.addWidget(self.add_course_button, 1, 1)
        class_area_layout.addLayout(class_buttons_layout)
        list_layout.addLayout(class_area_layout) # Sol tarafı ana list layout'una ekle

        # Öğretmen Listesi ve Butonları (Sağ Taraf)
        teacher_area_layout = QVBoxLayout()
        teacher_area_layout.addWidget(QLabel("Öğretmenler:"))
        self.teacher_list = QListWidget()
        self.teacher_list.itemClicked.connect(self.show_teacher_details)
        teacher_area_layout.addWidget(self.teacher_list)
        # Öğretmen Yönetim Butonları
        teacher_buttons_layout = QGridLayout() # Butonları grid ile dizelim
        self.add_teacher_button = QPushButton("Öğretmen Ekle")
        self.add_teacher_button.clicked.connect(self.add_teacher)
        teacher_buttons_layout.addWidget(self.add_teacher_button, 0, 0)
        self.update_teacher_button = QPushButton("Öğretmen Güncelle")
        self.update_teacher_button.clicked.connect(self.update_teacher)
        teacher_buttons_layout.addWidget(self.update_teacher_button, 0, 1)
        self.remove_teacher_button = QPushButton("Öğretmen Sil")
        self.remove_teacher_button.clicked.connect(self.remove_teacher)
        teacher_buttons_layout.addWidget(self.remove_teacher_button, 1, 0)
        self.assign_teacher_to_class_button = QPushButton("Sınıfa Öğrt. Ata") # İsmi kısalttık
        self.assign_teacher_to_class_button.clicked.connect(self.open_teacher_assignment_window)
        teacher_buttons_layout.addWidget(self.assign_teacher_to_class_button, 1, 1)
        teacher_area_layout.addLayout(teacher_buttons_layout)
        list_layout.addLayout(teacher_area_layout) # Sağ tarafı ana list layout'una ekle

        # Üst widget'ı ana layout'a ekle
        main_layout.addWidget(top_widget)
        # Listelerin olduğu layout'u ana layout'a ekle
        main_layout.addLayout(list_layout)

        # Alt Kısım (Hesaplama ve Program Oluşturma)
        bottom_layout = QHBoxLayout()
        self.calculate_hours_button = QPushButton("Toplam Ders Saatini Göster")
        self.calculate_hours_button.clicked.connect(self.show_total_lesson_summary)
        #self.calculate_hours_button.clicked.connect(self.calculate_total_hours)
        bottom_layout.addWidget(self.calculate_hours_button)
        


        # HİBRİT MODEL İÇİN YENİ GİRİŞ ALANLARI
        form_layout = QFormLayout()
        self.exploration_input = QLineEdit("1000")
        self.candidates_input = QLineEdit("10")
        form_layout.addRow("Keşif Denemesi Sayısı:", self.exploration_input)
        form_layout.addRow("İyileştirilecek Aday Sayısı:", self.candidates_input)
        
        bottom_layout.addLayout(form_layout)
        
        # <<< YENİ BUTON: Ders Programı Oluştur >>>
        self.generate_timetable_button = QPushButton("Ders Programı Oluştur")
        self.generate_timetable_button.clicked.connect(self.generate_timetable)
        bottom_layout.addWidget(self.generate_timetable_button)
        
        # CSP butonu (isteğe bağlı)
        #self.csp_button = QPushButton("CSP ile Oluştur (Deneysel)")
        #self.csp_button.clicked.connect(self.run_csp_scheduler)
        #bottom_layout.addWidget(self.csp_button)

        main_layout.addLayout(bottom_layout)
        # <<< SON >>>

        self.setLayout(main_layout) # Ana layout'u ayarla
        self.set_school_level(self.school_level_dropdown.currentText())


    # --- Diğer Metotlar (set_school_level, create_classes, select_class, ...) ---
    # Bu metotların çoğu önceki halindeki gibi kalabilir.
    # Sadece veri yapısı değişikliklerinden etkilenenler güncellenmişti.
    # (Kodun önceki mesajdaki hali güncel kabul ediliyor)

    def set_school_level(self, level):
        self.school_level = level
        for i in reversed(range(self.section_layout.count())):
            widget_item = self.section_layout.itemAt(i)
            if widget_item is not None:
                widget = widget_item.widget()
                if widget is not None:
                    widget.deleteLater()
        self.section_inputs.clear()
        if level == "Lise": grades = ["9", "10", "11", "12"]
        elif level == "Ortaokul": grades = ["5", "6", "7", "8"]
        else: grades = ["1", "2", "3", "4"] # İlkokul
        for i, grade in enumerate(grades):
            label = QLabel(f"{grade}. sınıf şube sayısı:")
            self.section_layout.addWidget(label, i, 0)
            input_field = QLineEdit()
            input_field.setPlaceholderText("Sayısal değer girin")
            self.section_layout.addWidget(input_field, i, 1)
            self.section_inputs[grade] = input_field
        # Okul kademesi değişince listeleri de temizle ve UI'ı yenile
        self.classes.clear()
        self.teachers.clear() # Öğretmenleri de temizlemek mantıklı mı? Opsiyonel.
        self.refresh_ui_lists()


    def create_classes(self):
        try:
            daily_hours_text = self.daily_hours_input.text().strip()
            if not daily_hours_text: raise ValueError("Günlük ders saati boş olamaz!")
            if not daily_hours_text.isdigit(): raise ValueError("Günlük ders saati sadece sayısal bir değer olmalıdır!")
            self.daily_hours = int(daily_hours_text)
            if self.daily_hours <= 0: raise ValueError("Günlük ders saati 0'dan büyük olmalıdır!")

            # Mevcut sınıfları korumak yerine temizleyip yeniden oluşturuyoruz.
            # Eğer mevcut sınıflara ekleme/çıkarma isteniyorsa mantık değişmeli.
            self.class_list.clear()
            self.classes.clear() # Temizle

            any_section_created = False
            grade_counts = {} # Hangi seviyeden kaç şube istendiğini tutalım

            # Önce tüm girdileri oku ve doğrula
            for grade, input_field in self.section_inputs.items():
                section_count_text = input_field.text().strip()
                if not section_count_text:
                    grade_counts[grade] = 0
                    continue
                if not section_count_text.isdigit(): raise ValueError(f"{grade}. sınıf şube sayısı sadece sayısal bir değer olmalıdır!")
                section_count = int(section_count_text)
                if section_count < 0: raise ValueError(f"{grade}. sınıf şube sayısı negatif olamaz!")
                grade_counts[grade] = section_count
                if section_count > 0: any_section_created = True

            if not any_section_created:
                 QMessageBox.warning(self, "Uyarı", "Hiç şube sayısı girilmedi veya 0 girildi. Sınıf oluşturulamadı.")
                 self.refresh_ui_lists() # Listeyi temizle
                 return

            # Doğrulama başarılıysa sınıfları oluştur
            for grade, section_count in grade_counts.items():
                 if section_count > 0:
                     for section in range(section_count):
                         class_name = f"{grade}{chr(65 + section)}"
                         # Aynı isimde sınıf varsa üzerine yazma (veya uyar)
                         if class_name not in self.classes:
                              self.classes[class_name] = {'courses': [], 'hours': {}, 'assigned_teachers': {}, 'distribution': {}}
                         else:
                              QMessageBox.warning(self, "Uyarı", f"'{class_name}' sınıfı zaten mevcut, üzerine yazılmayacak.")


            self.refresh_ui_lists() # UI listesini güncelle
            QMessageBox.information(self, "Başarı", "Girilen şube sayılarına göre sınıflar oluşturuldu/güncellendi!")

        except ValueError as e:
            QMessageBox.critical(self, "Hata", f"Geçersiz giriş: {str(e)}")
            # Hata durumunda listeyi temizlemek yerine mevcut durumu koruyabiliriz.
            # self.classes.clear() # Temizleme opsiyonel
            # self.refresh_ui_lists()
            self.daily_hours = 0 # Hatalı saat girdisi varsa sıfırla

    def select_class(self, item):
        self.selected_class = item.text()

    # assign_courses, save_courses, set_course_hours, save_course_hours metotları önceki mesajdaki güncel halleriyle kalabilir.

    def assign_courses(self):
        if not self.selected_class:
            QMessageBox.warning(self, "Hata", "Lütfen önce bir sınıf seçin!")
            return
        if self.selected_class not in self.classes:
            QMessageBox.warning(self, "Hata", f"'{self.selected_class}' sınıfı bulunamadı veya düzgün oluşturulmamış.")
            return

        self.course_window = QWidget()
        self.course_window.setWindowTitle(f"'{self.selected_class}' Sınıfına Ders Ata")
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content_widget = QWidget()
        scroll.setWidget(content_widget)
        layout = QVBoxLayout(content_widget)

        self.course_checkboxes = {}
        sorted_courses = sorted(list(set(self.courses))) # Genel ders listesi

        if not sorted_courses:
            layout.addWidget(QLabel("Henüz hiç ders eklenmemiş."))
        else:
            current_class_courses = self.classes[self.selected_class].get('courses', [])
            for course in sorted_courses:
                checkbox = QCheckBox(course)
                if course in current_class_courses:
                    checkbox.setChecked(True)
                layout.addWidget(checkbox)
                self.course_checkboxes[course] = checkbox

        main_layout = QVBoxLayout(self.course_window)
        main_layout.addWidget(scroll)
        save_button = QPushButton("Seçili Dersleri Kaydet")
        save_button.clicked.connect(self.save_courses)
        main_layout.addWidget(save_button)
        self.course_window.resize(400, 500)
        self.course_window.show()

    def save_courses(self):
        if not self.selected_class or self.selected_class not in self.classes:
            QMessageBox.critical(self, "Hata", "Sınıf seçimiyle ilgili bir sorun oluştu.")
            if hasattr(self, 'course_window'): self.course_window.close()
            return

        selected_courses = sorted([course for course, checkbox in self.course_checkboxes.items() if checkbox.isChecked()])
        current_class_data = self.classes[self.selected_class]
        current_courses = current_class_data.get('courses', [])
        current_hours = current_class_data.get('hours', {})
        current_teachers = current_class_data.get('assigned_teachers', {})

        new_hours = {}
        new_teachers = {}
        removed_courses_info = []

        for course in selected_courses:
            new_hours[course] = current_hours.get(course, 0)
            if course in current_teachers:
                new_teachers[course] = current_teachers[course]

        for course in current_courses:
            if course not in selected_courses:
                removed_info = f"'{course}'"
                if course in current_teachers:
                    teacher_id = current_teachers[course]
                    removed_info += f" ({teacher_id} öğretmeninden)"
                    if teacher_id in self.teachers and 'assignments' in self.teachers[teacher_id] and course in self.teachers[teacher_id]['assignments']:
                         if self.selected_class in self.teachers[teacher_id]['assignments'][course]:
                             self.teachers[teacher_id]['assignments'][course].remove(self.selected_class)
                             if not self.teachers[teacher_id]['assignments'][course]: del self.teachers[teacher_id]['assignments'][course]
                             if not self.teachers[teacher_id]['assignments']: del self.teachers[teacher_id]['assignments']
                             if hasattr(self, 'ta_window') and self.ta_window.isVisible():
                                 self.update_teacher_assignment_list_from_teachers()
                removed_courses_info.append(removed_info)

        self.classes[self.selected_class]['courses'] = selected_courses
        self.classes[self.selected_class]['hours'] = new_hours
        self.classes[self.selected_class]['assigned_teachers'] = new_teachers

        if removed_courses_info:
            QMessageBox.information(self, "Bilgi",
                                    f"'{self.selected_class}' sınıfından kaldırılan dersler:\n- " + "\n- ".join(removed_courses_info) +
                                    "\nİlgili öğretmen atamaları da kaldırıldı.")
        else:
             QMessageBox.information(self, "Başarı", f"'{self.selected_class}' sınıfının dersleri güncellendi!")

        if hasattr(self, 'course_window'): self.course_window.close()

    # LÜTFEN MEVCUT set_course_hours FONKSİYONUNU BU BLOK İLE DEĞİŞTİRİN
    # LÜTFEN MEVCUT set_course_hours FONKSİYONUNU BU BLOK İLE DEĞİŞTİRİN
    # LÜTFEN MEVCUT set_course_hours FONKSİYONUNU BU BLOK İLE DEĞİŞTİRİN
    def set_course_hours(self):
        if not self.selected_class:
            QMessageBox.warning(self, "Hata", "Lütfen önce bir sınıf seçin!")
            return
        class_data = self.classes.get(self.selected_class)
        if not isinstance(class_data, dict) or not class_data.get('courses'):
            QMessageBox.warning(self, "Hata", f"'{self.selected_class}' sınıfına henüz hiç ders atanmamış veya yapı bozuk.")
            return

        self.hours_window = QWidget()
        self.hours_window.setWindowTitle(f"'{self.selected_class}' Ders Saatlerini ve Dağılımını Belirle")
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content_widget = QWidget()
        scroll.setWidget(content_widget)
        
        layout = QGridLayout(content_widget)
        # Bu satır ders saati sütununun genişlemesini sağlar
        layout.setColumnStretch(1, 1) 
        
        # --- YENİ EKLENEN SATIR BURASI ---
        # Bu satır, dağılımın bulunduğu 3. sütunun (indeks 2)
        # pencere genişledikçe esnemesini ve büyümesini sağlar. Sorunu çözen satır budur.
        layout.setColumnStretch(2, 1)

        self.course_hours_inputs = {}
        self.course_distribution_labels = {}

        current_hours = class_data.get('hours', {})
        current_distributions = class_data.get('distribution', {})

        layout.addWidget(QLabel("<b>Ders Adı</b>"), 0, 0)
        layout.addWidget(QLabel("<b>Haftalık Saat</b>"), 0, 1)
        layout.addWidget(QLabel("<b>Dağılım (Bloklar)</b>"), 0, 2, 1, 3)

        for i, course in enumerate(sorted(class_data['courses'])):
            row = i + 1
            
            layout.addWidget(QLabel(course + ":"), row, 0)

            input_field = QLineEdit()
            current_hour = current_hours.get(course, 0)
            input_field.setText(str(current_hour))
            input_field.setPlaceholderText("Sayısal saat")
            layout.addWidget(input_field, row, 1)
            self.course_hours_inputs[course] = input_field

            distribution_layout = QHBoxLayout()
            
            distribution_label = QLabel()
            distribution_label.setMinimumWidth(80) # Bu komut doğru, kalmalı
            
            current_dist = current_distributions.get(course, [])
            distribution_label.setText(f"Dağılım: {','.join(map(str, current_dist))}" if current_dist else "Dağılım: Yok")
            distribution_layout.addWidget(distribution_label)
            self.course_distribution_labels[course] = distribution_label

            for num in range(1, 7):
                btn = QPushButton(str(num))
                btn.setFixedSize(25, 25)
                btn.clicked.connect(lambda checked, c=course, n=num: self._update_distribution_label(c, n))
                distribution_layout.addWidget(btn)

            clear_btn = QPushButton("Temizle")
            clear_btn.setFixedSize(60, 25)
            clear_btn.clicked.connect(lambda checked, c=course: self._clear_distribution_label(c))
            distribution_layout.addWidget(clear_btn)
            
            layout.addLayout(distribution_layout, row, 2)

        main_layout = QVBoxLayout(self.hours_window)
        main_layout.addWidget(scroll)
        save_button = QPushButton("Saatleri ve Dağılımları Kaydet")
        save_button.clicked.connect(self.save_course_hours)
        main_layout.addWidget(save_button)
        self.hours_window.resize(650, 500)
        self.hours_window.show()
    def _update_distribution_label(self, course, number):
        """Dağılım etiketine tıklandığında sayıyı ekler ve ders saatini kontrol eder."""
        # İlgili arayüz elemanlarını al
        label = self.course_distribution_labels.get(course)
        hours_input = self.course_hours_inputs.get(course)
        if not label or not hours_input: return

        # 1. Dersin toplam saatini al
        try:
            total_hours = int(hours_input.text())
            if total_hours <= 0:
                QMessageBox.warning(self.hours_window, "Uyarı", f"Lütfen önce '{course}' dersi için geçerli (0'dan büyük) bir saat girin.")
                return
        except (ValueError, TypeError):
            QMessageBox.warning(self.hours_window, "Uyarı", f"Lütfen önce '{course}' dersi için sayısal bir saat değeri girin.")
            return

        # 2. Mevcut dağılımın toplamını hesapla
        current_text = label.text().replace("Dağılım: ", "").replace("Yok", "")
        current_parts = [int(part) for part in current_text.split(',') if part.strip().isdigit()]
        current_sum = sum(current_parts)

        # 3. YENİ KONTROL: Yeni sayı eklenince toplam ders saati aşılıyor mu?
        if current_sum + number > total_hours:
            QMessageBox.warning(self.hours_window, "Geçersiz Dağılım",
                                f"'{course}' dersi için dağılım toplamı ({current_sum + number}), "
                                f"dersin toplam saatini ({total_hours}) aşamaz.")
            return # Aşıyorsa ekleme yapma

        # 4. Ekleme işlemini yap
        current_parts.append(number)
        label.setText(f"Dağılım: {','.join(map(str, current_parts))}")

    def _clear_distribution_label(self, course):
        """İlgili dersin dağılım etiketini temizler."""
        label = self.course_distribution_labels.get(course)
        if label:
            label.setText("Dağılım: Yok")

    # LÜTFEN MEVCUT save_course_hours FONKSİYONUNU BU BLOK İLE DEĞİŞTİRİN
    def save_course_hours(self):
        if not self.selected_class or self.selected_class not in self.classes:
            QMessageBox.critical(self, "Hata", "Sınıf seçimiyle ilgili bir sorun oluştu.")
            if hasattr(self, 'hours_window'): self.hours_window.close()
            return

        hours_changed = False
        invalid_input_found = False
        
        # 'hours' ve 'distribution' anahtarlarının var olduğundan emin ol
        current_class_hours = self.classes[self.selected_class].setdefault('hours', {})
        current_class_distributions = self.classes[self.selected_class].setdefault('distribution', {})

        for course, input_field in self.course_hours_inputs.items():
            try:
                # 1. Saatleri Kaydet (Mevcut mantık ile aynı)
                hours_text = input_field.text().strip()
                hours = 0
                if hours_text.isdigit():
                    hours = int(hours_text)
                    if hours < 0:
                        QMessageBox.warning(self.hours_window, "Geçersiz Giriş", f"'{course}' dersi için negatif saat girilemez.")
                        invalid_input_found = True
                        continue
                elif hours_text:
                    QMessageBox.warning(self.hours_window, "Geçersiz Giriş", f"'{course}' dersi için '{hours_text}' sayısal değil.")
                    invalid_input_found = True
                    continue

                if current_class_hours.get(course, -1) != hours:
                    current_class_hours[course] = hours
                    hours_changed = True
                
                # 2. YENİ: Dağılımları Kaydet
                distribution_label = self.course_distribution_labels.get(course)
                if distribution_label:
                    dist_text = distribution_label.text().replace("Dağılım: ", "").replace("Yok", "")
                    new_distribution = [int(p.strip()) for p in dist_text.split(',') if p.strip().isdigit()]
                    
                    # Doğrulama: Dağılım toplamı, ders saatine eşit mi?
                    if new_distribution and sum(new_distribution) != hours:
                        QMessageBox.warning(self.hours_window, "Uyumsuzluk",
                                        f"'{course}' dersi için girilen dağılımın toplamı ({sum(new_distribution)}), "
                                        f"ders saatine ({hours}) eşit değil.\nLütfen dağılımı düzeltin veya temizleyin.")
                        invalid_input_found = True
                        continue # Bu ders için kaydı atla, diğerlerine devam et

                    # Değişiklik varsa kaydet
                    if current_class_distributions.get(course, []) != new_distribution:
                        if new_distribution:
                            current_class_distributions[course] = new_distribution
                        elif course in current_class_distributions: # Dağılım silindiyse
                            del current_class_distributions[course]
                        hours_changed = True

            except Exception as e:
                QMessageBox.critical(self.hours_window, "Hata", f"'{course}' dersi kaydedilirken hata oluştu: {str(e)}")
                invalid_input_found = True

        if hours_changed and not invalid_input_found:
            QMessageBox.information(self, "Başarı", f"'{self.selected_class}' sınıfının ders saatleri ve dağılımları kaydedildi!")
        elif not hours_changed and not invalid_input_found:
            QMessageBox.information(self, "Bilgi", "Herhangi bir değişiklik yapılmadı.")

        if not invalid_input_found:
            if hasattr(self, 'hours_window'): self.hours_window.close()
    def add_course(self):
        course_name, ok = QInputDialog.getText(self, "Genel Ders Listesine Ekle", "Yeni dersin adını girin:")
        if ok and course_name:
            course_name = course_name.strip()
            if course_name:
                if course_name not in self.courses:
                    self.courses.append(course_name)
                    QMessageBox.information(self, "Başarı", f"'{course_name}' dersi genel listeye eklendi.")
                else:
                    QMessageBox.warning(self, "Uyarı", f"'{course_name}' dersi zaten listede mevcut.")
            else:
                QMessageBox.warning(self, "Uyarı", "Boş ders adı eklenemez.")


    # remove_course metodu önceki mesajdaki güncel haliyle kalabilir.
    def remove_course(self):
        if not self.selected_class:
            QMessageBox.warning(self, "Hata", "Lütfen önce bir sınıf seçin!")
            return

        class_data = self.classes.get(self.selected_class)
        if not isinstance(class_data, dict) or not class_data.get('courses'):
             QMessageBox.warning(self, "Hata", f"'{self.selected_class}' sınıfına henüz ders atanmamış veya yapı bozuk.")
             return

        assigned_courses = sorted(class_data['courses'])
        if not assigned_courses:
             QMessageBox.warning(self, "Hata", f"'{self.selected_class}' sınıfına henüz ders atanmamış.")
             return

        course_name, ok = QInputDialog.getItem(self, f"'{self.selected_class}' Sınıfından Ders Sil",
                                               "Silmek istediğiniz dersi seçin:",
                                               assigned_courses, 0, False)

        if ok and course_name:
            assigned_teacher = class_data.get('assigned_teachers', {}).get(course_name)
            teacher_info = f" ({assigned_teacher} öğretmeninden)" if assigned_teacher else ""

            reply = QMessageBox.question(self, 'Ders Silme Onayı',
                                         f"'{course_name}' dersini '{self.selected_class}' sınıfından silmek istediğinizden emin misiniz?\n"
                                         f"Varsa, ilgili ders saati ve öğretmen ataması{teacher_info} da kaldırılacaktır.",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return

            if course_name in class_data['courses']: class_data['courses'].remove(course_name)
            if course_name in class_data.get('hours', {}): del class_data['hours'][course_name]
            if course_name in class_data.get('distribution', {}): del class_data['distribution'][course_name]
            teacher_id_to_update = None
            if course_name in class_data.get('assigned_teachers', {}):
                 teacher_id_to_update = class_data['assigned_teachers'][course_name]
                 del class_data['assigned_teachers'][course_name]

            if teacher_id_to_update:
                 if teacher_id_to_update in self.teachers and 'assignments' in self.teachers[teacher_id_to_update] and course_name in self.teachers[teacher_id_to_update]['assignments']:
                     if self.selected_class in self.teachers[teacher_id_to_update]['assignments'][course_name]:
                         self.teachers[teacher_id_to_update]['assignments'][course_name].remove(self.selected_class)
                         if not self.teachers[teacher_id_to_update]['assignments'][course_name]: del self.teachers[teacher_id_to_update]['assignments'][course_name]
                         if not self.teachers[teacher_id_to_update]['assignments']: del self.teachers[teacher_id_to_update]['assignments']
                         if hasattr(self, 'ta_window') and self.ta_window.isVisible():
                             self.update_teacher_assignment_list_from_teachers()

            QMessageBox.information(self, "Başarı", f"'{course_name}' dersi '{self.selected_class}' sınıfından (ve ilgili atamalardan) silindi.")


    # calculate_total_hours metodu önceki mesajdaki güncel haliyle kalabilir.
    def calculate_total_hours(self):
        if not self.classes:
            QMessageBox.information(self, "Bilgi", "Henüz hiç sınıf oluşturulmamış.")
            return

        total_hours_all = 0
        subject_hours = {}
        classes_without_hours = []
        unassigned_lessons_count = 0

        for class_name, class_data in self.classes.items():
            class_total = 0
            hours_dict = class_data.get('hours', {})
            courses_in_class = class_data.get('courses', [])
            assigned_teachers_in_class = class_data.get('assigned_teachers', {})

            if not courses_in_class: continue # Ders atanmamış sınıfı atla

            has_hours = False
            for course in courses_in_class:
                 hours_assigned = hours_dict.get(course, 0)
                 if hours_assigned > 0:
                      has_hours = True
                      class_total += hours_assigned
                      subject_hours[course] = subject_hours.get(course, 0) + hours_assigned
                      # Öğretmen atanmış mı kontrolü
                      if course not in assigned_teachers_in_class:
                           unassigned_lessons_count += hours_assigned
                 elif course in hours_dict and hours_assigned == 0: # Saati 0 girilmişse
                     has_hours = True # Saat girilmiş sayılır

            total_hours_all += class_total
            if not has_hours and courses_in_class: # Dersi var ama saati hiç girilmemişse
                 classes_without_hours.append(class_name)


        if total_hours_all == 0 and not classes_without_hours:
            QMessageBox.information(self, "Bilgi", "Henüz hiçbir sınıfa ders saati atanmamış veya tüm saatler 0.")
            return

        result_message = f"Okulun Toplam Haftalık Ders Saati Yükü: {total_hours_all}\n\n"
        if subject_hours:
             result_message += "Branşlara Göre Toplam Saatler:\n"
             for subject, total in sorted(subject_hours.items()):
                 if total > 0:
                     result_message += f"- {subject}: {total} saat\n"
        else:
             result_message += "Henüz branşlara göre saat bilgisi girilmemiş.\n"

        if classes_without_hours:
            result_message += f"\nDersleri atanmış ancak saatleri girilmemiş/sıfır olan sınıflar: {', '.join(classes_without_hours)}\n"
        if unassigned_lessons_count > 0:
            result_message += f"\nDİKKAT: Sınıflara atanmış ancak öğretmeni atanmamış toplam {unassigned_lessons_count} ders saati bulunmaktadır!\n"


        QMessageBox.information(self, "Toplam Ders Saatleri", result_message)

    # add_teacher, update_teacher, _show_teacher_edit_window, _add_course_to_teacher_list,
    # _remove_course_from_teacher_list, save_teacher, remove_teacher, show_teacher_details
    # metotları önceki mesajdaki güncel halleriyle kalabilir.

    def add_teacher(self):
        if self.daily_hours <= 0:
             QMessageBox.warning(self, "Hata", "Öğretmen ekleyebilmek için önce geçerli bir 'Günlük Ders Saati' belirlemelisiniz!")
             return
        self._show_teacher_edit_window(is_update=False)


    def update_teacher(self):
        selected_items = self.teacher_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Hata", "Lütfen güncellemek istediğiniz öğretmeni listeden seçin!")
            return
        teacher_id = selected_items[0].text()
        if teacher_id not in self.teachers:
            QMessageBox.critical(self, "Hata", "Seçilen öğretmen verisi bulunamadı!")
            return
        if self.daily_hours <= 0:
            QMessageBox.warning(self, "Hata", "Öğretmen güncelleyebilmek için geçerli bir 'Günlük Ders Saati' belirlenmiş olmalı!")
            return
        self._show_teacher_edit_window(is_update=True, teacher_id=teacher_id)


    def _show_teacher_edit_window(self, is_update=False, teacher_id=None):
        # Bu fonksiyonda görsel bir değişiklik yapılmadı, grid layout zaten kullanılıyor.
        if is_update and (teacher_id is None or teacher_id not in self.teachers):
            QMessageBox.critical(self,"Hata", "Güncellenecek öğretmen bilgisi bulunamadı.")
            return
        teacher_data = self.teachers.get(teacher_id, {}) if is_update else {}

        self.teacher_window = QWidget()
        window_title = f"'{teacher_id}' Öğretmenini Güncelle" if is_update else "Yeni Öğretmen Ekle"
        self.teacher_window.setWindowTitle(window_title)
        layout = QVBoxLayout()

        name_layout = QHBoxLayout()
        self.teacher_name_input = QLineEdit(teacher_data.get('name', ''))
        self.teacher_name_input.setPlaceholderText("Ad")
        self.teacher_surname_input = QLineEdit(teacher_data.get('surname', ''))
        self.teacher_surname_input.setPlaceholderText("Soyad")
        name_layout.addWidget(self.teacher_name_input)
        name_layout.addWidget(self.teacher_surname_input)
        layout.addLayout(name_layout)

        course_layout = QHBoxLayout()
        self.teacher_course_selector = QComboBox()
        available_courses = sorted(list(set(self.courses)))
        if not available_courses:
            course_layout.addWidget(QLabel("Önce genel listeye ders eklemelisiniz."))
            self.teacher_course_selector.setEnabled(False)
            self.add_course_to_teacher_btn = QPushButton("Dersi Öğretmene Ekle", enabled=False)
        else:
            self.teacher_course_selector.addItems(available_courses)
            self.teacher_course_selector.setEnabled(True)
            self.add_course_to_teacher_btn = QPushButton("Dersi Öğretmene Ekle")
            self.add_course_to_teacher_btn.clicked.connect(self._add_course_to_teacher_list)

        course_layout.addWidget(self.teacher_course_selector)
        course_layout.addWidget(self.add_course_to_teacher_btn)
        layout.addLayout(course_layout)

        layout.addWidget(QLabel("Öğretmenin Verebileceği Dersler:"))
        self.assigned_courses_listwidget = QListWidget()
        if is_update:
            self.assigned_courses_listwidget.addItems(teacher_data.get('courses', []))

        layout.addWidget(self.assigned_courses_listwidget)
        self.remove_course_from_teacher_btn = QPushButton("Seçili Dersi Kaldır")
        self.remove_course_from_teacher_btn.clicked.connect(self._remove_course_from_teacher_list)
        layout.addWidget(self.remove_course_from_teacher_btn)

        availability_label = QLabel("Müsait Olmadığı Saatler (İşaretliler = Derse Giremez):")
        layout.addWidget(availability_label)
        scroll_availability = QScrollArea()
        scroll_availability.setWidgetResizable(True)
        grid_widget = QWidget()
        self.availability_grid_layout = QGridLayout(grid_widget)
        days = self.days_of_week # ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
        for col, day in enumerate(days):
            self.availability_grid_layout.addWidget(QLabel(day), 0, col + 1, Qt.AlignCenter)

        self.availability_checkboxes = {}
        if self.daily_hours > 0 :
            for row in range(self.daily_hours):
                self.availability_grid_layout.addWidget(QLabel(f"{row + 1}. Ders"), row + 1, 0, Qt.AlignRight)
                self.availability_checkboxes[row+1] = {}
                for col, day in enumerate(days):
                    cb = QCheckBox()
                    teacher_availability = teacher_data.get('availability', {})
                    if day in teacher_availability and (row + 1) in teacher_availability.get(day,[]):
                         cb.setChecked(True) # İşaretli = Müsait Değil

                    self.availability_grid_layout.addWidget(cb, row + 1, col + 1, Qt.AlignCenter)
                    self.availability_checkboxes[row+1][day] = cb
            scroll_availability.setWidget(grid_widget)
            layout.addWidget(scroll_availability)
        else:
             layout.addWidget(QLabel("Müsaitlik durumu için önce günlük ders saatini belirleyin."))


        action_text = "Güncelle" if is_update else "Kaydet"
        save_btn = QPushButton(action_text)
        save_btn.clicked.connect(lambda checked, update=is_update, old_id=teacher_id: self.save_teacher(update=update, old_id=old_id))
        layout.addWidget(save_btn)

        self.teacher_window.setLayout(layout)
        self.teacher_window.resize(600, 550)
        self.teacher_window.show()

    def _add_course_to_teacher_list(self):
        if not hasattr(self, 'teacher_course_selector') or not self.teacher_course_selector.isEnabled(): return
        course = self.teacher_course_selector.currentText()
        if course and not self.assigned_courses_listwidget.findItems(course, Qt.MatchExactly):
            self.assigned_courses_listwidget.addItem(course)

    def _remove_course_from_teacher_list(self):
        if not hasattr(self, 'assigned_courses_listwidget'): return
        selected_items = self.assigned_courses_listwidget.selectedItems()
        if selected_items:
            # Sadece listeden kaldırır, atamalara dokunmaz.
            self.assigned_courses_listwidget.takeItem(self.assigned_courses_listwidget.row(selected_items[0]))


    def save_teacher(self, update=False, old_id=None):
        if not all(hasattr(self, attr) for attr in ['teacher_name_input', 'teacher_surname_input', 'assigned_courses_listwidget', 'teacher_window']):
             QMessageBox.critical(self, "Hata", "Öğretmen bilgileri kaydedilirken bir sorun oluştu (Eksik Arayüz Elemanı).")
             return
        name = self.teacher_name_input.text().strip()
        surname = self.teacher_surname_input.text().strip()
        if not name or not surname:
            QMessageBox.warning(self.teacher_window, "Hata", "Öğretmen adı ve soyadı boş bırakılamaz!")
            return
        new_teacher_id = f"{name} {surname}"

        if not update and new_teacher_id in self.teachers:
            QMessageBox.warning(self.teacher_window, "Hata", f"'{new_teacher_id}' adında bir öğretmen zaten mevcut!")
            return
        if update and old_id != new_teacher_id and new_teacher_id in self.teachers:
            QMessageBox.warning(self.teacher_window, "Hata", f"'{new_teacher_id}' adında başka bir öğretmen zaten mevcut! İsim değişikliği yapılamıyor.")
            return

        assigned_courses = sorted([self.assigned_courses_listwidget.item(i).text()
                             for i in range(self.assigned_courses_listwidget.count())])
        if not assigned_courses:
            QMessageBox.warning(self.teacher_window, "Hata", "Öğretmene en az bir ders atanmalıdır (verebileceği ders)!")
            return

        availability = {}
        days = self.days_of_week
        if self.daily_hours > 0 and hasattr(self, 'availability_checkboxes'): # Checkboxlar varsa oku
            for day in days:
                unavailable_hours = []
                for hour in range(1, self.daily_hours + 1):
                    # Checkbox var mı diye kontrol et (dinamik olarak oluşturulduğu için)
                    if hour in self.availability_checkboxes and day in self.availability_checkboxes.get(hour,{}):
                         if self.availability_checkboxes[hour][day].isChecked():
                             unavailable_hours.append(hour)
                if unavailable_hours:
                    availability[day] = sorted(unavailable_hours)


        current_assignments = {}
        if update and old_id in self.teachers:
            current_assignments = self.teachers[old_id].get('assignments', {})
            # Güncelleme sırasında öğretmenin verebileceği derslerden çıkarılanlar olduysa,
            # ilgili dersin sınıf atamalarını da temizlemek gerekir mi? Bu kritik bir karar.
            # Şimdilik temizlemiyoruz, sadece 'courses' listesi güncelleniyor.
            # İstenirse, kaldırılan derslerin atamaları da temizlenebilir.
            removed_courses_from_list = set(current_assignments.keys()) - set(assigned_courses)
            if removed_courses_from_list:
                 print(f"Uyarı: Öğretmen {new_teacher_id} için kaldırılan dersler: {removed_courses_from_list}. Bu derslere ait sınıf atamaları korunuyor.")
                 # İstenirse buraya atama temizleme kodu eklenebilir.


        teacher_data = {
            'name': name, 'surname': surname, 'courses': assigned_courses,
            'availability': availability, 'assignments': current_assignments
        }

        id_changed = update and old_id and old_id != new_teacher_id

        if id_changed:
             if old_id in self.teachers: del self.teachers[old_id]
             items = self.teacher_list.findItems(old_id, Qt.MatchExactly)
             if items: self.teacher_list.takeItem(self.teacher_list.row(items[0]))
             self.teachers[new_teacher_id] = teacher_data
             self.teacher_list.addItem(new_teacher_id)
             # Sınıflardaki ID'yi güncelle
             for class_name, class_data in self.classes.items():
                  if isinstance(class_data, dict) and 'assigned_teachers' in class_data:
                      teachers_to_update = class_data['assigned_teachers']
                      courses_to_update = list(teachers_to_update.keys())
                      for course in courses_to_update:
                          if teachers_to_update.get(course) == old_id:
                               teachers_to_update[course] = new_teacher_id
             QMessageBox.information(self.teacher_window, "Başarılı", f"Öğretmen ID'si '{old_id}' -> '{new_teacher_id}' olarak güncellendi!")

        elif update and old_id:
             self.teachers[old_id] = teacher_data
             QMessageBox.information(self.teacher_window, "Başarılı", f"'{old_id}' öğretmeni başarıyla güncellendi!")
        else: # Yeni ekleme
             self.teachers[new_teacher_id] = teacher_data
             self.teacher_list.addItem(new_teacher_id)
             QMessageBox.information(self.teacher_window, "Başarılı", f"'{new_teacher_id}' öğretmeni başarıyla eklendi!")

        self.teacher_list.sortItems() # Öğretmen listesini sırala
        if hasattr(self, 'teacher_window'): self.teacher_window.close()
        self.refresh_ui_lists() # Diğer UI elemanlarını güncelle


    def remove_teacher(self):
        selected_items = self.teacher_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Hata", "Lütfen silmek istediğiniz öğretmeni listeden seçin!")
            return
        teacher_id = selected_items[0].text()

        reply = QMessageBox.question(self, 'Öğretmen Silme Onayı',
                                     f"'{teacher_id}' öğretmenini silmek istediğinizden emin misiniz?\n"
                                     f"Bu öğretmenin TÜM sınıf atamaları da kaldırılacaktır.",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            teacher_data_to_delete = self.teachers.get(teacher_id)

            if teacher_id in self.teachers: del self.teachers[teacher_id]

            items = self.teacher_list.findItems(teacher_id, Qt.MatchExactly)
            if items: self.teacher_list.takeItem(self.teacher_list.row(items[0]))

            cleaned_count = 0
            if teacher_data_to_delete and 'assignments' in teacher_data_to_delete:
                for course, classes in teacher_data_to_delete['assignments'].items():
                    for class_name in classes:
                         if class_name in self.classes and isinstance(self.classes[class_name], dict) and \
                            'assigned_teachers' in self.classes[class_name] and \
                            course in self.classes[class_name]['assigned_teachers'] and \
                            self.classes[class_name]['assigned_teachers'][course] == teacher_id:
                                 del self.classes[class_name]['assigned_teachers'][course]
                                 cleaned_count += 1

            QMessageBox.information(self, "Başarı", f"'{teacher_id}' öğretmeni ve ilgili {cleaned_count} sınıf ataması silindi.")
            self.refresh_ui_lists() # UI'ı yenile


    def show_teacher_details(self, item):
        teacher_id = item.text()
        if teacher_id in self.teachers:
            data = self.teachers[teacher_id]
            details = f"Ad: {data.get('name', 'N/A')}\n"
            details += f"Soyad: {data.get('surname', 'N/A')}\n"
            details += f"Verebileceği Dersler: {', '.join(data.get('courses', []))}\n\n"

            details += "Müsait Olmadığı Zamanlar:\n"
            availability = data.get('availability', {})
            if availability:
                days = self.days_of_week
                for day in days:
                     hours = availability.get(day, [])
                     if hours:
                          details += f"  {day}: {', '.join(map(str, hours))}. saatler\n"
                if not any(availability.get(d) for d in days): # Hiçbir gün dolu değilse
                     details += "  (Belirtilen bir zaman yok)\n"

            else:
                details += "  (Her zaman müsait görünüyor)\n"


            details += "\nAtandığı Sınıflar/Dersler:\n"
            assignments = data.get('assignments', {})
            if assignments:
                 for course, classes in sorted(assignments.items()):
                      if classes:
                           details += f"  {course}: {', '.join(sorted(classes))}\n"
            else:
                 details += "  (Henüz bir sınıfa atanmamış)\n"

            QMessageBox.information(self, f"Öğretmen Detayları: {teacher_id}", details)
        else:
            QMessageBox.warning(self, "Hata", "Öğretmen bilgisi bulunamadı.")


    # --- Sınıfa Öğretmen Atama Metotları ---
    # open_teacher_assignment_window, update_teacher_assignment_course_dropdown,
    # add_teacher_assignment, update_teacher_assignment_list_from_teachers,
    # remove_selected_assignment_entry metotları önceki mesajdaki güncel halleriyle kalabilir.
    def open_teacher_assignment_window(self):
        self.ta_window = QWidget()
        self.ta_window.setWindowTitle("Sınıfa Öğretmen Ata")
        layout = QVBoxLayout()

        form_layout = QFormLayout()

        self.ta_teacher_dropdown = QComboBox()
        self.ta_teacher_dropdown.addItem("Yok")
        self.ta_teacher_dropdown.addItems(sorted(self.teachers.keys()))
        self.ta_teacher_dropdown.currentTextChanged.connect(self.update_teacher_assignment_course_dropdown)
        form_layout.addRow("Öğretmen:", self.ta_teacher_dropdown)

        self.ta_course_dropdown = QComboBox()
        self.ta_course_dropdown.addItem("Yok")
        form_layout.addRow("Ders:", self.ta_course_dropdown)

        self.ta_class_dropdown = QComboBox()
        self.ta_class_dropdown.addItem("Yok")
        self.ta_class_dropdown.addItems(sorted(self.classes.keys()))
        form_layout.addRow("Sınıf:", self.ta_class_dropdown)

        layout.addLayout(form_layout)

        add_btn = QPushButton("Seçili Atamayı Ekle")
        add_btn.clicked.connect(self.add_teacher_assignment)
        layout.addWidget(add_btn)

        self.assignment_list_widget = QListWidget()
        layout.addWidget(QLabel("Mevcut Atamalar:"))
        layout.addWidget(self.assignment_list_widget)

        remove_btn = QPushButton("Seçili Atamayı (Tüm Sınıflarıyla) Kaldır")
        remove_btn.clicked.connect(self.remove_selected_assignment_entry)
        layout.addWidget(remove_btn)

        self.ta_window.setLayout(layout)
        self.ta_window.resize(550, 450)
        self.update_teacher_assignment_list_from_teachers()
        self.update_teacher_assignment_course_dropdown(self.ta_teacher_dropdown.currentText())
        self.ta_window.show()

    def update_teacher_assignment_course_dropdown(self, teacher_id):
        self.ta_course_dropdown.clear()
        self.ta_course_dropdown.addItem("Yok")
        if teacher_id != "Yok" and teacher_id in self.teachers:
            teacher_data = self.teachers.get(teacher_id, {})
            assigned_courses = teacher_data.get('courses', [])
            if assigned_courses:
                self.ta_course_dropdown.addItems(sorted(assigned_courses))

    def add_teacher_assignment(self):
        teacher_id = self.ta_teacher_dropdown.currentText()
        course = self.ta_course_dropdown.currentText()
        class_name = self.ta_class_dropdown.currentText()

        if teacher_id == "Yok" or course == "Yok" or class_name == "Yok":
            QMessageBox.warning(self.ta_window, "Uyarı", "Lütfen öğretmen, ders ve sınıf için geçerli bir seçim yapın.")
            return

        # Öğretmen ve Sınıf var mı kontrolü
        if teacher_id not in self.teachers:
            QMessageBox.critical(self.ta_window, "Hata", f"Öğretmen '{teacher_id}' bulunamadı!")
            return
        if class_name not in self.classes or not isinstance(self.classes[class_name], dict):
             QMessageBox.critical(self.ta_window, "Hata", f"Sınıf '{class_name}' bulunamadı veya yapısı bozuk!")
             return

        # Öğretmenin o dersi verebilir mi kontrolü (opsiyonel ama önerilir)
        if course not in self.teachers[teacher_id].get('courses', []):
             QMessageBox.warning(self.ta_window, "Uyarı", f"Öğretmen '{teacher_id}' normalde '{course}' dersini vermiyor.\nYine de atamak istiyorsanız devam edebilirsiniz, ancak öğretmenin ders listesini güncellemeniz önerilir.")
             # İsterseniz burada return ile işlemi durdurabilirsiniz.

        # Sınıfın o dersi aldığından emin ol (yoksa ekle)
        class_courses = self.classes[class_name].setdefault('courses', [])
        if course not in class_courses:
             reply = QMessageBox.question(self.ta_window, 'Onay',
                                         f"'{course}' dersi normalde '{class_name}' sınıfının ders listesinde yok.\n"
                                         f"Derse öğretmen atanırken sınıfın ders listesine de eklensin mi?",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
             if reply == QMessageBox.Yes:
                 class_courses.append(course)
                 class_courses.sort()
                 # Ders saati 0 olarak eklenebilir, sonra ayarlanmalı
                 self.classes[class_name].setdefault('hours', {})[course] = self.classes[class_name]['hours'].get(course, 0)

             else:
                  QMessageBox.information(self.ta_window,"İptal", "Sınıfın ders listesi güncellenmediği için atama yapılmadı.")
                  return


        # Çakışma Kontrolü ve Atama
        class_teachers = self.classes[class_name].setdefault('assigned_teachers', {})
        current_teacher = class_teachers.get(course)

        if current_teacher and current_teacher != teacher_id:
            QMessageBox.warning(self.ta_window, "Çakışma!",
                                f"'{class_name}' sınıfındaki '{course}' dersi zaten '{current_teacher}' öğretmenine atanmış.\n"
                                f"'{teacher_id}' atanamadı. Önce mevcut atamayı kaldırın.")
            return
        elif current_teacher == teacher_id:
             QMessageBox.information(self.ta_window, "Bilgi", f"Bu atama zaten mevcut.")
             return
        else: # Atama yapılabilir
             # 1. Sınıfa ata
             class_teachers[course] = teacher_id

             # 2. Öğretmene ata
             self.teachers[teacher_id].setdefault('assignments', {})
             self.teachers[teacher_id]['assignments'].setdefault(course, [])
             if class_name not in self.teachers[teacher_id]['assignments'][course]:
                 self.teachers[teacher_id]['assignments'][course].append(class_name)
                 self.teachers[teacher_id]['assignments'][course].sort()

             QMessageBox.information(self.ta_window,"Başarılı",f"'{teacher_id}' öğretmeni '{course}' dersi için '{class_name}' sınıfına atandı.")
             self.update_teacher_assignment_list_from_teachers()


    def update_teacher_assignment_list_from_teachers(self):
        if not hasattr(self, 'assignment_list_widget'): return # Pencere kapalıysa işlem yapma
        self.assignment_list_widget.clear()
        for teacher_id, teacher_data in sorted(self.teachers.items()):
            if 'assignments' in teacher_data:
                for course, classes in sorted(teacher_data['assignments'].items()):
                    if classes:
                        classes_str = ", ".join(sorted(classes))
                        item_text = f"{teacher_id} {course}: {classes_str}"
                        self.assignment_list_widget.addItem(item_text)

    def remove_selected_assignment_entry(self):
         if not hasattr(self, 'assignment_list_widget'): return
         selected_items = self.assignment_list_widget.selectedItems()
         if not selected_items:
             QMessageBox.warning(self.ta_window, "Uyarı", "Lütfen kaldırmak istediğiniz atamayı listeden seçin.")
             return

         item = selected_items[0]
         text = item.text()

         reply = QMessageBox.question(self.ta_window, 'Onay',
                                      f"'{text}'\n\nYukarıdaki öğretmen-ders atamasını TÜM sınıflarından kaldırmak istediğinizden emin misiniz?",
                                      QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

         if reply == QMessageBox.Yes:
             try:
                 parts = text.split(':')
                 teacher_course_part = parts[0].strip()
                 last_space_index = teacher_course_part.rfind(' ')
                 if last_space_index == -1: raise ValueError("Format hatası.")
                 teacher_id = teacher_course_part[:last_space_index]
                 course = teacher_course_part[last_space_index+1:]

                 assigned_classes = []
                 if teacher_id in self.teachers and 'assignments' in self.teachers[teacher_id] and course in self.teachers[teacher_id]['assignments']:
                      assigned_classes = self.teachers[teacher_id]['assignments'][course][:]

                 if not assigned_classes:
                      # Belki liste güncel değil? Yine de öğretmenden silmeyi dene.
                      if teacher_id in self.teachers and 'assignments' in self.teachers[teacher_id] and course in self.teachers[teacher_id]['assignments']:
                           del self.teachers[teacher_id]['assignments'][course]
                           if not self.teachers[teacher_id]['assignments']: del self.teachers[teacher_id]['assignments']
                           QMessageBox.warning(self.ta_window, "Bilgi", f"Öğretmen kaydında sınıf bulunamadı ama '{teacher_id} - {course}' ataması kaldırıldı.")
                           self.update_teacher_assignment_list_from_teachers()
                      else:
                           QMessageBox.warning(self.ta_window, "Hata", "Atama verisi bulunamadı.")
                      return

                 teacher_updated = False
                 if teacher_id in self.teachers and 'assignments' in self.teachers[teacher_id] and course in self.teachers[teacher_id]['assignments']:
                     del self.teachers[teacher_id]['assignments'][course]
                     if not self.teachers[teacher_id]['assignments']: del self.teachers[teacher_id]['assignments']
                     teacher_updated = True

                 classes_updated_count = 0
                 for class_name in assigned_classes:
                      if class_name in self.classes and isinstance(self.classes[class_name], dict) and \
                         'assigned_teachers' in self.classes[class_name] and \
                         course in self.classes[class_name]['assigned_teachers'] and \
                         self.classes[class_name]['assigned_teachers'][course] == teacher_id:
                              del self.classes[class_name]['assigned_teachers'][course]
                              classes_updated_count += 1

                 if teacher_updated:
                      self.update_teacher_assignment_list_from_teachers()
                      QMessageBox.information(self.ta_window, "Başarılı", f"'{teacher_id} - {course}' ataması ({classes_updated_count} sınıftan) kaldırıldı.")
                 else:
                      QMessageBox.warning(self.ta_window, "Hata", "Atama öğretmenin kaydında bulunamadı veya zaten silinmiş.")

             except ValueError as ve:
                  QMessageBox.warning(self.ta_window, "Format Hatası", f"Atama metni ayrıştırılamadı: {ve}")
             except Exception as e:
                 QMessageBox.critical(self.ta_window, "Hata", f"Atama kaldırılırken beklenmedik bir hata oluştu: {str(e)}")


    # --- Dosya İşlemleri (generate_template, load_from_file, save_all_data, load_saved_data) ---
    # Bu metotlar önceki mesajdaki güncel halleriyle kalabilir.
    def generate_template(self):
        template_content = """
# Okul Ders Programı Veri Giriş Şablonu
# Lütfen ilgili alanları doldurun ve '#' ile başlayan satırları silin veya yorum olarak bırakın.

Okul Kademesi: Lise # Seçenekler: Lise, Ortaokul, İlkokul
Günlük Ders Saati: 8 # Günlük toplam ders saati sayısı

# Sınıf Seviyeleri ve Şube Sayıları (İlgili kademeye göre doldurun)
9. Sınıf Şube Sayısı: 2
10. Sınıf Şube Sayısı: 1
# ...

# --- Sınıf Bilgileri ---
SINIF: 9A
Ders: Türk Dili ve Edebiyatı, Saat: 5, Öğretmen: Ayşe Yılmaz
Ders: Matematik, Saat: 6, Öğretmen: Ali Veli
Ders: Yabancı Dil (İngilizce), Saat: 4, Öğretmen: Serkan Bıçakcı
# ... diğer dersler

SINIF: 9B
# ... 9B'nin dersleri ...

SINIF: 10A
Ders: Seçmeli İngilizce, Saat: 2, Öğretmen: Serkan Bıçakcı
# ... 10A'nın diğer dersleri ...

# --- Öğretmen Bilgileri ---
ÖĞRETMEN: Serkan Bıçakcı
Ad: Serkan
Soyad: Bıçakcı
Dersler: Yabancı Dil (İngilizce), Seçmeli İngilizce
Müsait Değil: Pazartesi 1, Pazartesi 2
Müsait Değil: Cuma 8

ÖĞRETMEN: Ayşe Yılmaz
Ad: Ayşe
Soyad: Yılmaz
Dersler: Türk Dili ve Edebiyatı, Dil Bilgisi

ÖĞRETMEN: Ali Veli
Ad: Ali
Soyad: Veli
Dersler: Matematik, Temel Matematik, İleri Matematik

# ... diğer öğretmenler ...
"""
        self.show_text_in_window("Veri Giriş Şablonu", template_content)


    def load_from_file(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Notepad Dosyasını Seç", "",
                                                   "Text Files (*.txt);;All Files (*)", options=options)
        if not file_path: return

        try:
            with open(file_path, 'r', encoding='utf-8') as f: lines = f.readlines()

            temp_classes = {}
            temp_teachers = {}
            temp_courses = list(self.courses) # Mevcutları koru, yenileri ekle
            temp_daily_hours = 0
            temp_school_level = self.school_level
            temp_section_counts = {}

            current_context = None
            current_class_name = None
            current_teacher_id = None

            print("Dosya okunuyor...") # Debug
            for i, line in enumerate(lines):
                line = line.strip()
                if not line or line.startswith('#'): continue

                # print(f"Satır {i+1}: {line}") # Debug

                if line.startswith("Okul Kademesi:"): temp_school_level = line.split(":", 1)[1].strip()
                elif line.startswith("Günlük Ders Saati:"):
                    try: temp_daily_hours = int(line.split(":", 1)[1].strip())
                    except ValueError: print(f"Uyarı: Geçersiz günlük saat formatı: {line}")
                elif ". Sınıf Şube Sayısı:" in line:
                     try:
                         parts = line.split(". Sınıf Şube Sayısı:")
                         grade = parts[0].strip()
                         count = int(parts[1].strip())
                         temp_section_counts[grade] = count
                     except: print(f"Uyarı: Geçersiz şube formatı: {line}")

                elif line.startswith("SINIF:"):
                    current_class_name = line.split(":", 1)[1].strip()
                    if current_class_name:
                        temp_classes[current_class_name] = {'courses': [], 'hours': {}, 'assigned_teachers': {}}
                        current_context = "SINIF"; current_teacher_id = None
                        # print(f"-> Sınıf: {current_class_name}") # Debug
                    else: current_context = None
                elif line.startswith("ÖĞRETMEN:"):
                    current_teacher_id = line.split(":", 1)[1].strip()
                    if current_teacher_id:
                        temp_teachers[current_teacher_id] = {'name': '', 'surname': '', 'courses': [], 'availability': {}, 'assignments': {}}
                        current_context = "ÖĞRETMEN"; current_class_name = None
                        # print(f"-> Öğretmen: {current_teacher_id}") # Debug
                    else: current_context = None

                elif current_context == "SINIF" and current_class_name:
                    if line.startswith("Ders:"):
                        try:
                             parts = line.split(',')
                             course_part = parts[0].split(":", 1)[1].strip()
                             if course_part not in temp_classes[current_class_name]['courses']:
                                  temp_classes[current_class_name]['courses'].append(course_part)
                             if course_part not in temp_courses: temp_courses.append(course_part)

                             hour = 0; teacher = None
                             for part in parts[1:]:
                                 part = part.strip()
                                 if part.startswith("Saat:"): hour = int(part.split(":", 1)[1].strip())
                                 elif part.startswith("Öğretmen:"): teacher = part.split(":", 1)[1].strip()

                             if hour > 0: temp_classes[current_class_name]['hours'][course_part] = hour
                             if teacher:
                                  temp_classes[current_class_name]['assigned_teachers'][course_part] = teacher
                                  if teacher not in temp_teachers: # Öğretmen henüz tanımlanmadıysa ekle
                                       t_name = teacher.split()[0] if ' ' in teacher else teacher
                                       t_surname = teacher.split()[-1] if ' ' in teacher else ''
                                       temp_teachers[teacher] = {'name': t_name, 'surname': t_surname, 'courses': [], 'availability': {}, 'assignments': {}}
                                  # Öğretmenin atamasını da ekle
                                  temp_teachers[teacher].setdefault('assignments', {})
                                  temp_teachers[teacher]['assignments'].setdefault(course_part, [])
                                  if current_class_name not in temp_teachers[teacher]['assignments'][course_part]:
                                       temp_teachers[teacher]['assignments'][course_part].append(current_class_name)

                        except Exception as e: print(f"Uyarı: Sınıf ders format hatası '{current_class_name}': {line} -> {e}")

                elif current_context == "ÖĞRETMEN" and current_teacher_id:
                    if line.startswith("Ad:"): temp_teachers[current_teacher_id]['name'] = line.split(":", 1)[1].strip()
                    elif line.startswith("Soyad:"):
                         temp_teachers[current_teacher_id]['surname'] = line.split(":", 1)[1].strip()
                         name = temp_teachers[current_teacher_id]['name']
                         surname = temp_teachers[current_teacher_id]['surname']
                         full_name = f"{name} {surname}"
                         if full_name != current_teacher_id and name and surname:
                             if full_name not in temp_teachers:
                                 temp_teachers[full_name] = temp_teachers.pop(current_teacher_id)
                                 # Sınıflardaki eski ID'yi güncellemek GEREKİR! (Okuma sonrası yapılmalı)
                                 current_teacher_id = full_name # ID'yi takip et
                             else: print(f"Uyarı: Öğretmen ID çakışması '{current_teacher_id}' vs '{full_name}'")

                    elif line.startswith("Dersler:"):
                        courses = [c.strip() for c in line.split(":", 1)[1].split(',') if c.strip()]
                        temp_teachers[current_teacher_id]['courses'] = courses
                        for c in courses:
                             if c not in temp_courses: temp_courses.append(c)
                    elif line.startswith("Müsait Değil:"):
                         try:
                              parts = line.split(":", 1)[1].split()
                              day = parts[0]
                              hours = [int(h.replace(',','').strip()) for h in parts[1:] if h.replace(',','').strip().isdigit()]
                              if day in self.days_of_week and hours:
                                   temp_teachers[current_teacher_id]['availability'].setdefault(day, [])
                                   temp_teachers[current_teacher_id]['availability'][day].extend(hours)
                                   temp_teachers[current_teacher_id]['availability'][day] = sorted(list(set(temp_teachers[current_teacher_id]['availability'][day])))
                         except Exception as e: print(f"Uyarı: Müsait değil format hatası '{current_teacher_id}': {line} -> {e}")

            # <<< ID Güncelleme Kontrolü (Dosya okuma sonrası) >>>
            id_map = {} # old_id -> new_id
            teachers_copy = list(temp_teachers.items()) # Üzerinde iterate ederken değiştirmek için kopya
            for t_id, data in teachers_copy:
                 name = data.get('name','')
                 surname = data.get('surname','')
                 if name and surname:
                      full_name = f"{name} {surname}"
                      if t_id != full_name:
                           if full_name not in temp_teachers: # Yeni ID boşta mı?
                                temp_teachers[full_name] = temp_teachers.pop(t_id)
                                id_map[t_id] = full_name
                           # else: Çakışma varsa eski ID kalır (yukarıda uyarı verildi)

            # Sınıflardaki öğretmen ID'lerini güncelle
            if id_map:
                 print("Öğretmen ID'leri güncelleniyor...")
                 for class_name, class_data in temp_classes.items():
                      assigned_tch = class_data.get('assigned_teachers',{})
                      courses_to_update = list(assigned_tch.keys())
                      for course in courses_to_update:
                           old_tid = assigned_tch.get(course)
                           if old_tid in id_map:
                                assigned_tch[course] = id_map[old_tid] # Yeni ID ile değiştir

            # Her şey yolundaysa ana değişkenleri güncelle
            self.school_level = temp_school_level
            self.daily_hours = temp_daily_hours
            self.courses = sorted(list(set(temp_courses)))
            self.classes = temp_classes
            self.teachers = temp_teachers

            # UI'ı Güncelle
            index = self.school_level_dropdown.findText(self.school_level)
            if index != -1: self.school_level_dropdown.setCurrentIndex(index)
            self.daily_hours_input.setText(str(self.daily_hours))
            self.set_school_level(self.school_level) # Bu section_inputs'u ayarlar
            # Okunan şube sayılarını UI'a yazdır
            for grade, count in temp_section_counts.items():
                 if grade in self.section_inputs:
                      self.section_inputs[grade].setText(str(count))

            self.refresh_ui_lists()

            QMessageBox.information(self, "Başarılı", "Veriler dosyadan başarıyla yüklendi.\n'Şubeleri Oluştur' ile sınıf listesini, öğretmen bilgileriyle öğretmen listesini kontrol edebilirsiniz.")

        except FileNotFoundError: QMessageBox.critical(self, "Hata", "Dosya bulunamadı!")
        except Exception as e: QMessageBox.critical(self, "Dosya Okuma Hatası", f"Dosya okunurken bir hata oluştu: {str(e)}")


    def save_all_data(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(self, "Verileri Kaydet", "",
                                                   "JSON Files (*.json);;All Files (*)", options=options)
        if not file_path: return
        if not file_path.lower().endswith('.json'): file_path += '.json'

        data_to_save = {
            'school_level': self.school_level,
            'daily_hours': self.daily_hours,
            'courses': self.courses,
            'classes': self.classes,
            'teachers': self.teachers,
        }
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data_to_save, f, ensure_ascii=False, indent=4)
            QMessageBox.information(self, "Başarılı", f"Tüm veriler başarıyla '{os.path.basename(file_path)}' dosyasına kaydedildi.")
        except Exception as e:
            QMessageBox.critical(self, "Kaydetme Hatası", f"Veriler kaydedilirken bir hata oluştu: {str(e)}")

    def load_saved_data(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Kayıtlı Verileri Yükle", "",
            "JSON Files (*.json);;All Files (*)", options=options
            )
        if not file_path:
            return
    
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                loaded_data = json.load(f)
    
            # Verileri yükle
            self.school_level = loaded_data.get('school_level', 'Lise')
            self.daily_hours = loaded_data.get('daily_hours', 0)
            self.courses = loaded_data.get('courses', [])
            self.classes = loaded_data.get('classes', {})
            self.teachers = loaded_data.get('teachers', {})
    
            # UI ELEMENTLERİNİ TAMAMEN YENİLE
            # 1. Okul kademesini ayarla
            index = self.school_level_dropdown.findText(self.school_level)
            if index >= 0:
                self.school_level_dropdown.setCurrentIndex(index)
            
            # 2. Günlük ders saatini ayarla
            self.daily_hours_input.setText(str(self.daily_hours))
            
            # 3. Sınıf şube inputlarını temizle (yeniden oluşturmak için)
            for grade, input_field in self.section_inputs.items():
                input_field.clear()
            
            # 4. Sınıf ve öğretmen listelerini güncelle
            self.refresh_ui_lists()
            
            # 5. Öğretmen atama penceresi açıksa kapat
            if hasattr(self, 'ta_window') and self.ta_window.isVisible():
                self.ta_window.close()
    
            QMessageBox.information(
                self, "Başarılı",
                f"Veriler başarıyla yüklendi!\n\n"
                f"• Yüklenen sınıf sayısı: {len(self.classes)}\n"
                f"• Yüklenen öğretmen sayısı: {len(self.teachers)}\n\n"
                f"Artık sınıf listesini ve öğretmen listesini görebilirsiniz.\n"
                f"Lütfen 'Şubeleri Oluştur' butonuna BASMAYIN!"
            )
    
        except Exception as e:
            QMessageBox.critical(
                self, "Hata",
                f"Veri yüklenirken hata oluştu:\n{str(e)}"
            )
    def refresh_ui_lists(self):
        # Sınıf listesini temizle ve yeniden doldur
        self.class_list.clear()
        self.class_list.addItems(sorted(self.classes.keys()))
        
        # Öğretmen listesini temizle ve yeniden doldur
        self.teacher_list.clear()
        self.teacher_list.addItems(sorted(self.teachers.keys()))
    
        # Öğretmen atama penceresi varsa onu da güncelle
        if hasattr(self, 'ta_window') and self.ta_window.isVisible():
            current_teacher = self.ta_teacher_dropdown.currentText()
            self.ta_teacher_dropdown.clear()
            self.ta_teacher_dropdown.addItem("Yok")
            self.ta_teacher_dropdown.addItems(sorted(self.teachers.keys()))
            
            idx = self.ta_teacher_dropdown.findText(current_teacher)
            if idx != -1: 
                self.ta_teacher_dropdown.setCurrentIndex(idx)
            
            current_class = self.ta_class_dropdown.currentText()
            self.ta_class_dropdown.clear()
            self.ta_class_dropdown.addItem("Yok")
            self.ta_class_dropdown.addItems(sorted(self.classes.keys()))
            
            idx = self.ta_class_dropdown.findText(current_class)
            if idx != -1: 
                self.ta_class_dropdown.setCurrentIndex(idx)
            
                self.update_teacher_assignment_course_dropdown(self.ta_teacher_dropdown.currentText())
                self.update_teacher_assignment_list_from_teachers()


    def show_text_in_window(self, title, text):
        text_window = QWidget()
        text_window.setWindowTitle(title)
        layout = QVBoxLayout()
        text_edit = QTextEdit()
        text_edit.setPlainText(text)
        text_edit.setReadOnly(True)
        layout.addWidget(text_edit)
        close_button = QPushButton("Kapat")
        close_button.clicked.connect(text_window.close)
        layout.addWidget(close_button)
        text_window.setLayout(layout)
        text_window.resize(600, 700)
        setattr(self, f"_{title.replace(' ','_').lower()}_window", text_window)
        text_window.show()


    # <<< ============================================= >>>
    # <<<        DERS PROGRAMI OLUŞTURMA BÖLÜMÜ        >>>
    # <<< ============================================= >>>
    def _place_lesson_safely(self, teacher, course, class_name, day, hour, class_timetables, teacher_timetables):
        # Bu fonksiyonun tutarlılığı garanti ettiği varsayılır. Diğer algoritmalarda da aynı yöntem kullanılmalı.
        lesson_info = {
            "teacher": teacher,
            "course": course,
            "class": class_name
        }
        class_timetables[class_name][day][hour] = lesson_info
        teacher_timetables[teacher][day][hour] = lesson_info


    def generate_timetable(self):
        """
        Hibrit Model ile çalışan ana ders programı oluşturma fonksiyonu.
        Aşama 1: Geniş keşif ile en iyi adayları bulur.
        Aşama 2: Sadece en iyi adaylar üzerinde yoğun kurtarma operasyonu çalıştırır.
        """
        if not self._check_data_readiness():
            return

        try:
            exploration_attempts = int(self.exploration_input.text() or "1000")
            candidates_to_optimize = int(self.candidates_input.text() or "10")
            if exploration_attempts < 1 or candidates_to_optimize < 1: raise ValueError()
        except (ValueError, TypeError):
            QMessageBox.warning(self, "Geçersiz Giriş", "Lütfen geçerli sayısal değerler girin.")
            return

        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Eksik Kütüphane", "'openpyxl' kütüphanesi kurulu değil.")
            return

        # 1. Yerleştirilecek tüm dersleri hazırla
        all_lessons_to_place = self._get_lessons_to_place()
        all_lessons_to_place.sort(key=lambda l: self.calculate_difficulty_score(l['teacher'], l['course'], l['class']), reverse=True)
        
        # --- AŞAMA 1: GENİŞ KEŞİF ---
        progress1 = QProgressDialog("Aşama 1: En iyi çözümler keşfediliyor...", "İptal", 0, exploration_attempts, self)
        progress1.setWindowTitle("Lütfen Bekleyin (Keşif Aşaması)")
        progress1.setWindowModality(Qt.WindowModal)
        
        all_results = []
        for attempt in range(exploration_attempts):
            progress1.setValue(attempt)
            if progress1.wasCanceled():
                return
            QApplication.processEvents()

            class_timetables = {cn: [[None for _ in range(self.daily_hours)] for _ in range(len(self.days_of_week))] for cn in self.classes}
            teacher_timetables = {tn: [[None for _ in range(self.daily_hours)] for _ in range(len(self.days_of_week))] for tn in self.teachers}
            
            lessons_for_this_attempt = list(all_lessons_to_place)
            #slice_point = len(lessons_for_this_attempt) // 4
            #top_slice = lessons_for_this_attempt[:slice_point]
            #random.shuffle(top_slice)
            #lessons_for_this_attempt[:slice_point] = top_slice

            unplaced_lessons = []
            for lesson in lessons_for_this_attempt:
                placed_hours = 0
                hours_needed = lesson['hours']
                unplaced_blocks_for_this_lesson = []

                # Öncelik 1: Kullanıcı tanımlı dağılım varsa onu kullan
                if lesson['distribution']:
                    blocks_to_place = list(lesson['distribution'])
                    random.shuffle(blocks_to_place) # Blokları rastgele sırada yerleştirmeyi dene
                    
                    for block_size in blocks_to_place:
                        if self._find_and_place_lesson(lesson['teacher'], lesson['course'], lesson['class'], block_size, class_timetables, teacher_timetables):
                            placed_hours += block_size
                        else:
                            unplaced_blocks_for_this_lesson.append(block_size)
                
                # Öncelik 2: Dağılım yoksa, eski yöntemle (2'li ve 1'li bloklar) yerleştir
                else:
                    num_of_blocks = hours_needed // 2
                    for _ in range(num_of_blocks):
                        if self._find_and_place_lesson(lesson['teacher'], lesson['course'], lesson['class'], 2, class_timetables, teacher_timetables):
                            placed_hours += 2
                        else:
                            unplaced_blocks_for_this_lesson.append(2)

                    remaining_hours = hours_needed - placed_hours
                    for _ in range(remaining_hours):
                        if self._find_and_place_lesson(lesson['teacher'], lesson['course'], lesson['class'], 1, class_timetables, teacher_timetables):
                            placed_hours += 1
                        else:
                            unplaced_blocks_for_this_lesson.append(1)

                # Yerleşemeyen saatleri/blokları kaydet
                if placed_hours < hours_needed:
                    # Yeni yapı: Hangi blokların yerleşemediğini not alıyoruz.
                    # Bu, sonraki optimizasyon adımları için daha değerli bilgi sağlar.
                    # Şimdilik basitçe kalan saat sayısını tutmaya devam edelim.
                    unplaced_total_hours = sum(unplaced_blocks_for_this_lesson)
                    if unplaced_total_hours > 0:
                        unplaced_lessons.append({'teacher': lesson['teacher'], 'course': lesson['course'], 'class': lesson['class'], 'hours': unplaced_total_hours})

            unplaced_hours_count = sum(l['hours'] for l in unplaced_lessons)
            all_results.append({
                'unplaced_count': unplaced_hours_count,
                'class_tables': class_timetables,
                'teacher_tables': teacher_timetables,
                'unplaced_list': unplaced_lessons
            })

        progress1.close()
        progress1.deleteLater()

        if not all_results:
            QMessageBox.warning(self, "Hata", "Hiçbir çözüm üretilemedi.")
            return

        # En iyi N adayı seç
        all_results.sort(key=lambda x: x['unplaced_count'])
        top_candidates = all_results[:candidates_to_optimize]

        # --- AŞAMA 2: YOĞUN BAKIM (OPTİMİZASYON) ---
        progress2 = QProgressDialog("Aşama 2: En iyi adaylar iyileştiriliyor...", "İptal", 0, len(top_candidates), self)
        progress2.setWindowTitle("Lütfen Bekleyin (İyileştirme Aşaması)")
        progress2.setWindowModality(Qt.WindowModal)
        progress2.show()
        QApplication.processEvents()


        best_overall_result = None
        best_overall_unplaced_count = float('inf')

        for i, candidate in enumerate(top_candidates):
            progress2.setValue(i)
            progress2.setLabelText(f"Aday {i+1}/{len(top_candidates)} iyileştiriliyor... (Mevcut en iyi: {best_overall_unplaced_count} saat)")
            if progress2.wasCanceled():
                break
            
            # Kopyalar üzerinde çalışarak orijinal adayı bozmuyoruz
            class_tables_copy = {k: [row[:] for row in v] for k, v in candidate['class_tables'].items()}
            teacher_tables_copy = {k: [row[:] for row in v] for k, v in candidate['teacher_tables'].items()}
            unplaced_copy = [l.copy() for l in candidate['unplaced_list']]
            
            # İteratif kurtarma operasyonunu çalıştır
            rescue_attempts = 5
            for _ in range(rescue_attempts):
                if not unplaced_copy: break
                was_placed = self._force_place_unplaced_lessons(unplaced_copy, class_tables_copy, teacher_tables_copy)
                if not was_placed: break

            # Sonucu kontrol et ve genel en iyi ile karşılaştır
            current_unplaced_count = sum(l['hours'] for l in unplaced_copy)
            if current_unplaced_count < best_overall_unplaced_count:
                best_overall_unplaced_count = current_unplaced_count
                best_overall_result = (class_tables_copy, teacher_tables_copy, unplaced_copy)
            
            # Mükemmel çözümü bulduysak daha fazla bekleme
            if best_overall_unplaced_count == 0:
                break
        
        progress2.close()
        progress2.deleteLater()
        

        # --- FİNAL: En İyi Sonucu Raporla ---
        if best_overall_result:
            final_class_tables, final_teacher_tables, final_unplaced = best_overall_result

            # Eksik dersleri doğru say
            actual_unplaced_hours = sum(l['hours'] for l in final_unplaced if l['hours'] > 0)
            still_unplaced = [l for l in final_unplaced if l['hours'] > 0]
            
            print("Yerleşemeyen Dersler Listesi:")
            for l in still_unplaced:
                print(l)

            print(f"🔍 Gerçek Yerleştirilemeyen Saat: {actual_unplaced_hours}")
            print(f"🔍 Yerleşemeyen Ders Sayısı: {len(still_unplaced)}")

            # Gerekliyse SA başlat
            if still_unplaced:
                print("✅ Simulated Annealing başlatılıyor...")

                improved_class_tables, improved_teacher_tables, improved_unplaced = self._run_simulated_annealing(
                    final_class_tables, final_teacher_tables, final_unplaced
                )

                improved_unplaced_count = sum(l['hours'] for l in improved_unplaced if l['hours'] > 0)
                print("✅ Simulated Annealing tamamlandı.")
                print(f"⏬ Yeni Eksik Saat: {improved_unplaced_count}")

                if improved_unplaced_count < actual_unplaced_hours:
                    final_class_tables = improved_class_tables
                    final_teacher_tables = improved_teacher_tables
                    final_unplaced = improved_unplaced
                    actual_unplaced_hours = improved_unplaced_count

            # Rapor ve kayıt
            self.class_timetables = final_class_tables
            self.teacher_timetables = final_teacher_tables
            self.last_run_unplaced_lessons = final_unplaced
            total_hours = sum(l['hours'] for l in all_lessons_to_place)
            placed_count = total_hours - actual_unplaced_hours
            stats = {'total': total_hours, 'placed': placed_count}
            self._report_and_save_timetable(final_class_tables, final_teacher_tables, final_unplaced, stats)

            self.validate_schedule_consistency(final_class_tables, final_teacher_tables)
            self._check_bidirectional_consistency(final_class_tables, final_teacher_tables)
            self._sanitize_timetable(final_class_tables)
            self._sanitize_timetable(final_teacher_tables)

        else:
            QMessageBox.warning(self, "Hata", "İyileştirme aşamasında bir sorun oluştu.")
            return  # ← Bu satır çok önemli!


    def _run_simulated_annealing(self, class_tables, teacher_tables, unplaced_lessons, 
                           max_iterations=1000, initial_temp=100.0, cooling_rate=0.003):
        """Gelişmiş Simulated Annealing ile ders yerleştirme optimizasyonu"""
        import math
        import random
        import copy

        # 1. Başlangıç durumu
        current_class = {k: [row[:] for row in v] for k, v in class_tables.items()}
        current_teacher = {k: [row[:] for row in v] for k, v in teacher_tables.items()}
        current_unplaced = [l.copy() for l in unplaced_lessons]
        
        best_class = copy.deepcopy(current_class)
        best_teacher = copy.deepcopy(current_teacher)
        best_unplaced = copy.deepcopy(current_unplaced)
        best_cost = sum(l['hours'] for l in best_unplaced)

        # 2. Isı (temperature) parametreleri
        temp = initial_temp
        min_temp = 0.1
        
        for iteration in range(max_iterations):
            if best_cost == 0 or temp < min_temp:
                break

            # 3. Yeni komşu çözüm üret
            new_class, new_teacher, new_unplaced = self._generate_neighbor_solution(
                current_class, current_teacher, current_unplaced
            )
            
            # 4. Maliyet hesapla
            current_cost = sum(l['hours'] for l in current_unplaced)
            new_cost = sum(l['hours'] for l in new_unplaced)
            cost_diff = new_cost - current_cost

            # 5. Kabul kriteri
            if cost_diff < 0 or math.exp(-cost_diff / temp) > random.random():
                current_class, current_teacher, current_unplaced = new_class, new_teacher, new_unplaced
                current_cost = new_cost
                
                # En iyi çözümü güncelle
                if new_cost < best_cost:
                    best_class, best_teacher, best_unplaced = copy.deepcopy(new_class), copy.deepcopy(new_teacher), copy.deepcopy(new_unplaced)
                    best_cost = new_cost

            # 6. Soğutma
            temp *= (1 - cooling_rate)

        # 7. Sonuçları doğrula
        self._validate_improvement(unplaced_lessons, best_unplaced)
        return best_class, best_teacher, best_unplaced

    # Lütfen bu fonksiyonu kopyalayıp eskisinin üzerine yapıştırın
    def _generate_neighbor_solution(self, class_tables, teacher_tables, unplaced_lessons):
        """
        Komşu çözüm üretmek için daha etkili stratejiler kullanır. (GÜVENLİ VERSİYON)
        """
        new_class = {k: [row[:] for row in v] for k, v in class_tables.items()}
        new_teacher = {k: [row[:] for row in v] for k, v in teacher_tables.items()}
        new_unplaced = [l.copy() for l in unplaced_lessons]

        if new_unplaced and random.random() < 0.7:
            strategy = "place_unplaced"
        else:
            strategy = random.choice(["move", "swap"])
        
        if strategy == "place_unplaced" and new_unplaced:
            lesson_to_try = random.choice(new_unplaced)
            was_placed = self._aggressively_try_to_place(lesson_to_try, new_class, new_teacher) # Bu zaten güvenli olmalı
            if was_placed:
                for item in new_unplaced:
                    if item['teacher'] == lesson_to_try['teacher'] and item['class'] == lesson_to_try['class'] and item['course'] == lesson_to_try['course']:
                        item['hours'] -= 1
                        if item['hours'] <= 0:
                            new_unplaced.remove(item)
                        break
        
        elif strategy == "move":
            class_name = random.choice(list(new_class.keys()))
            day_idx, hour_idx = random.randint(0, len(self.days_of_week)-1), random.randint(0, self.daily_hours-1)
            
            lesson_to_move = new_class[class_name][day_idx][hour_idx]
            if isinstance(lesson_to_move, dict):
                # Önce GÜVENLİ BİR ŞEKİLDE SİL
                self._clear_lesson_safely(lesson_to_move, day_idx, hour_idx, new_class, new_teacher)
                
                new_day, new_hour = self._find_empty_slot_for_lesson(lesson_to_move['teacher'], lesson_to_move['course'], lesson_to_move['class'], new_class, new_teacher)
                
                if new_day is not None:
                    # Sonra GÜVENLİ BİR ŞEKİLDE YERLEŞTİR
                    self._place_lesson_safely(lesson_to_move['teacher'], lesson_to_move['course'], lesson_to_move['class'], new_day, new_hour, new_class, new_teacher)
                else:
                    # Taşıyamadıysak eski yerine GÜVENLİ BİR ŞEKİLDE geri koy
                    self._place_lesson_safely(lesson_to_move['teacher'], lesson_to_move['course'], lesson_to_move['class'], day_idx, hour_idx, new_class, new_teacher)

        elif strategy == "swap":
            class_name = random.choice(list(new_class.keys()))
            class_schedule = new_class[class_name]
            
            filled_slots = [(d, h) for d in range(len(self.days_of_week)) for h in range(self.daily_hours) if isinstance(class_schedule[d][h], dict)]
            
            if len(filled_slots) >= 2:
                (d1, h1), (d2, h2) = random.sample(filled_slots, 2)
                lesson1, lesson2 = class_schedule[d1][h1], class_schedule[d2][h2]

                # (Uygunluk kontrolleri burada)
                teacher1, teacher2 = lesson1['teacher'], lesson2['teacher']
                t1_available_at_slot2 = (h2 + 1) not in self.teachers[teacher1].get('availability', {}).get(self.days_of_week[d2], [])
                t2_available_at_slot1 = (h1 + 1) not in self.teachers[teacher2].get('availability', {}).get(self.days_of_week[d1], [])
                
                # Öğretmenlerin hedef slotları başka dersle dolu mu?
                t1_target_slot_ok = (new_teacher[teacher1][d2][h2] is None)
                t2_target_slot_ok = (new_teacher[teacher2][d1][h1] is None)

                if t1_available_at_slot2 and t2_available_at_slot1 and t1_target_slot_ok and t2_target_slot_ok:
                    # Her şey uygunsa, SWAP işlemini GÜVENLİ yap
                    # 1. Önce her iki dersi de GÜVENLİ SİL
                    self._clear_lesson_safely(lesson1, d1, h1, new_class, new_teacher)
                    self._clear_lesson_safely(lesson2, d2, h2, new_class, new_teacher)
                    
                    # 2. Sonra çaprazlama şekilde GÜVENLİ YERLEŞTİR
                    self._place_lesson_safely(teacher1, lesson1['course'], class_name, d2, h2, new_class, new_teacher)
                    self._place_lesson_safely(teacher2, lesson2['course'], class_name, d1, h1, new_class, new_teacher)

        return new_class, new_teacher, new_unplaced

    def _validate_improvement(self, original_unplaced, improved_unplaced):
        """İyileştirmeyi doğrular ve tutarsızlıkları loglar"""
        original_count = sum(l['hours'] for l in original_unplaced)
        improved_count = sum(l['hours'] for l in improved_unplaced)
        
        if improved_count > original_count:
            print(f"⚠️ Uyarı: Simulated annealing sonrası kötüleşme! (Önce: {original_count}, Sonra: {improved_count})")
        else:
            print(f"✅ Simulated annealing başarısı: {original_count - improved_count} saat kazanıldı")



    def _check_bidirectional_consistency(self, class_tables, teacher_tables):
        issues = []
        for class_name, table in class_tables.items():
            for day_idx, day in enumerate(table):
                for hour_idx, lesson in enumerate(day):
                    if isinstance(lesson, dict):
                        teacher = lesson.get("teacher")
                        course = lesson.get("course")
                        reverse = teacher_tables.get(teacher, [[]])[day_idx][hour_idx]
                        if not isinstance(reverse, dict) or (reverse is not lesson and reverse != lesson):
                            issues.append(
                                f"UYUMSUZLUK: {class_name} sınıfı {self.days_of_week[day_idx]} {hour_idx+1}. saat → {lesson}, "
                                f"ama {teacher} tablosunda bu saat {reverse}"
                            )
        if issues:
            print(f"🔴 TERS YÖNLÜ UYUŞMA SORUNLARI TESPİT EDİLDİ: {len(issues)} adet.")
            for issue in issues:
                print(issue)
        else:
            print("✅ Sınıf ve öğretmen tabloları tam uyumlu.")

    def test_simulated_annealing():
        # 1. Örnek veri hazırla
        class_tables = {...}  # Mevcut program
        teacher_tables = {...}
        unplaced = [...]  # Yerleştirilemeyen dersler
        
        # 2. Önceki durumu kaydet
        prev_unplaced_count = sum(l['hours'] for l in unplaced)
        
        # 3. Simulated annealing çalıştır
        new_class, new_teacher, new_unplaced = scheduler._run_simulated_annealing(
            class_tables, teacher_tables, unplaced
        )
        
        # 4. Sonuçları karşılaştır
        new_count = sum(l['hours'] for l in new_unplaced)
        assert new_count <= prev_unplaced_count, "İyileşme olmalıydı!"
        
        # 5. Tutarlılık kontrolü
        scheduler.validate_schedule_consistency(new_class, new_teacher)

    def _try_place_lesson(self, lesson, class_tables, teacher_tables):
        """Dersi yerleştirmeyi dener ve başarı durumunu döner"""
        placed = False
        for _ in range(lesson['hours']):
            placed |= self._find_and_place_lesson(
                lesson['teacher'], 
                lesson['course'], 
                lesson['class'], 
                1,  # 1 saatlik deneme
                class_tables, 
                teacher_tables
            )
        return placed
    

    def _free_up_space_for_lesson(self, lesson, class_timetables, teacher_timetables):
        """Zorunlu dersler için yer açmaya çalış"""
        teacher_id = lesson['teacher']
        course = lesson['course']
        class_name = lesson['class']
        days = self.days_of_week
        
        # 1. Öğretmenin en az dolu olan gününü bul
        best_day = None
        min_lessons = float('inf')
        
        for day_idx in range(len(days)):
            teacher_lessons = sum(1 for h in range(self.daily_hours) 
                                if teacher_timetables[teacher_id][day_idx][h] is not None)
            
            if teacher_lessons < min_lessons:
                min_lessons = teacher_lessons
                best_day = day_idx
        
        # 2. Bu günde bir dersi başka yere taşımayı dene
        if best_day is not None:
            for hour in range(self.daily_hours):
                if teacher_timetables[teacher_id][best_day][hour] is not None:
                    # Taşınacak ders bilgilerini al
                    moved_course_info = teacher_timetables[teacher_id][best_day][hour]
                    moved_class = moved_course_info.split('\n(')[-1].rstrip(')')
                    
                    # Dersi geçici olarak kaldır
                    teacher_timetables[teacher_id][best_day][hour] = None
                    class_timetables[moved_class][best_day][hour] = None
                    
                    # Yeni yer bulmaya çalış
                    success = self._find_and_place_lesson(
                        teacher_id, course, class_name, 1,  # 1 saatlik deneme
                        class_timetables, teacher_timetables
                    )
                    
                    if success:
                        # Yerleştirildi, şimdi taşınan dersi yeni yerine koy
                        moved_success = self._find_and_place_lesson(
                            teacher_id, moved_course_info.split('\n')[0], moved_class, 1,
                            class_timetables, teacher_timetables
                        )
                        if moved_success:
                            return True
                        else:
                            # Taşıma başarısız oldu, eski haline getir
                            teacher_timetables[teacher_id][best_day][hour] = moved_course_info
                            class_timetables[moved_class][best_day][hour] = f"{moved_course_info.split('\n')[0]}\n({teacher_id.split()[0]})"
                    else:
                        # Yerleştirme başarısız, eski haline getir
                        teacher_timetables[teacher_id][best_day][hour] = moved_course_info
                        class_timetables[moved_class][best_day][hour] = f"{moved_course_info.split('\n')[0]}\n({teacher_id.split()[0]})"
        
        return False




        # 4. Sonuçları Bildir ve Kaydet
    def _report_and_save_timetable(self, class_timetables, teacher_timetables, unplaced_lessons, stats):
        """Oluşturulan programı kullanıcıya bildirir ve Excel'e kaydeder."""
        total_lessons = stats['total']
        placed_lessons = stats['placed']
        unplaced_count = len(unplaced_lessons)
        success_rate = (placed_lessons / total_lessons * 100) if total_lessons > 0 else 0
    
        result_message = f"Ders Programı Oluşturma Tamamlandı!\n\n"
        result_message += f"Toplam Ders Saati: {total_lessons}\n"
        result_message += f"Yerleştirilen Ders Saati: {placed_lessons}\n"
        result_message += f"Başarı Oranı: {success_rate:.2f}%\n\n"
    
        if unplaced_lessons:
            result_message += f"UYARI: Aşağıdaki {unplaced_count} ders/saat yerleştirilemedi:\n"
            for item in unplaced_lessons[:10]: # İlk 10 tanesini göster
                result_message += f"- {item['class']} Sınıfı, {item['course']} Dersi ({item['teacher']}), {item['hours']} saat\n"
            if unplaced_count > 10: 
                result_message += "- ... (ve diğerleri)\n"
            result_message += "\nOluşturulan programda boşluklar olacaktır.\n"
        else:
            result_message += "Tüm dersler başarıyla yerleştirildi!\n"
    
        # Artık her durumda kaydetme sorusu sorulacak
        result_message += "\nProgramı Excel dosyası olarak kaydetmek ister misiniz?"
        
        reply = QMessageBox.information(self, "Program Oluşturma Sonucu", result_message,
                                        QMessageBox.Save | QMessageBox.Cancel, QMessageBox.Save)
    
        if reply == QMessageBox.Save:
            self._save_timetables_to_excel(class_timetables, teacher_timetables)
    def calculate_difficulty_score(self, teacher_id, course, class_name):
        """Bir dersin yerleştirilme zorluğunu puanlayan daha gelişmiş bir fonksiyon."""
        teacher_data = self.teachers.get(teacher_id, {})
        class_data = self.classes.get(class_name, {})
        if not teacher_data or not class_data:
            return 0

        score = 0
        
        # 1. Öğretmenin Esnekliği: Öğretmenin meşgul olduğu her saat, puanı artırır.
        availability = teacher_data.get('availability', {})
        unavailable_hours = sum(len(hours) for hours in availability.values())
        score += unavailable_hours * 5  # Meşgul her saat için 5 puan

        # 2. Dersin Büyüklüğü: Haftalık ders saati arttıkça yerleştirmek zorlaşır.
        course_hours = class_data.get('hours', {}).get(course, 0)
        score += course_hours * 10  # Haftalık her saat için 10 puan

        # 3. Blok Ders İhtiyacı: 2 saatten fazla olan dersler blok yerleştirme gerektirebileceğinden daha zordur.
        if course_hours > 2:
            score += 20 # Ekstra zorluk puanı
            
        # 4. Öğretmenin Toplam Yükü: Çok dersi olan öğretmenlerin programı daha sıkışıktır.
        total_teacher_load = 0
        if 'assignments' in teacher_data:
             for crs, c_list in teacher_data['assignments'].items():
                for c_name in c_list:
                    if c_name in self.classes:
                        total_teacher_load += self.classes[c_name].get('hours', {}).get(crs, 0)
        score += total_teacher_load * 2 # Toplam yükün 2 katı kadar puan

        # 5. Sınıfın Toplam Ders Yükü: Çok dersi olan bir sınıfa yeni ders eklemek daha zordur.
        total_class_hours = sum(class_data.get('hours', {}).values())
        score += total_class_hours
        
        return score
    def _find_empty_slot_for_lesson(self, teacher_id, course, class_name, class_timetables, teacher_timetables):
        """
        Belirli bir ders için programda tamamen boş olan ilk uygun yeri (gün, saat) bulur.
        Bu fonksiyon, yerinden oynatılan bir derse yeni bir yuva bulmak için kullanılır.
        """
        # Günleri ve saatleri rastgele deneyerek her seferinde farklı bir çözüm arayalım
        shuffled_day_indices = list(range(len(self.days_of_week)))
        random.shuffle(shuffled_day_indices)
        
        shuffled_hour_indices = list(range(self.daily_hours))
        random.shuffle(shuffled_hour_indices)

        for day_idx in shuffled_day_indices:
            day_name = self.days_of_week[day_idx]
            unavailable_hours_for_day = self.teachers[teacher_id].get('availability', {}).get(day_name, [])
            
            for hour_idx in shuffled_hour_indices:
                hour_human = hour_idx + 1 # 1-based index
                
                # Öğretmen müsait mi ve hem öğretmen hem de sınıf programı o saatte boş mu?
                if (hour_human not in unavailable_hours_for_day and
                        teacher_timetables[teacher_id][day_idx][hour_idx] is None and
                        class_timetables[class_name][day_idx][hour_idx] is None):
                    return day_idx, hour_idx # Uygun boş yer bulundu

        return None, None # Uygun boş yer bulunamadı

    def _force_place_unplaced_lessons(self, unplaced_lessons, class_timetables, teacher_timetables):
        random.shuffle(unplaced_lessons)
        progress_made = False

        for lesson in unplaced_lessons[:]:
            teacher = lesson['teacher']
            course = lesson['course']
            class_name = lesson['class']
            hours = lesson['hours']

            for _ in range(hours):
                placed = False
                for day in range(len(self.days_of_week)):
                    for hour in range(self.daily_hours):
                        if class_timetables[class_name][day][hour] is None and \
                        teacher_timetables[teacher][day][hour] is None:
                            self._place_lesson_safely(teacher, course, class_name, day, hour, class_timetables, teacher_timetables)
                            print(f"DOĞRUDAN YERLEŞTİRME: '{course}' ({class_name}) → {self.days_of_week[day]} {hour+1}. saat")
                            lesson['hours'] -= 1
                            if lesson['hours'] == 0:
                                unplaced_lessons.remove(lesson)
                            progress_made = True
                            placed = True
                            break
                    if placed:
                        break

        return progress_made
    
    
    def _aggressively_try_to_place(self, lesson_to_place, class_timetables, teacher_timetables):
        """
        Bir dersi, gerekirse başka bir dersi yerinden oynatarak ("tahliye ederek") yerleştirmeye çalışır.
        Sadece 1 saatlik bir bloğu yerleştirmeyi dener.
        """
        teacher = lesson_to_place['teacher']
        course = lesson_to_place['course']
        class_name = lesson_to_place['class']
        
        # Öğretmenin ve sınıfın tüm zaman dilimlerini potansiyel hedef olarak değerlendir
        shuffled_days = list(range(len(self.days_of_week)))
        random.shuffle(shuffled_days)
        shuffled_hours = list(range(self.daily_hours))
        random.shuffle(shuffled_hours)

        for day_idx in shuffled_days:
            day_name = self.days_of_week[day_idx]
            # Öğretmen o gün o saatte müsait değilse (kendi takviminden dolayı), o yuvayı hiç deneme
            if (shuffled_hours[0] + 1) in self.teachers[teacher].get('availability', {}).get(day_name, []):
                continue

            for hour_idx in shuffled_hours:
                # 1. HEDEF YUVAYI BELİRLE
                target_slot_class = class_timetables[class_name][day_idx][hour_idx]
                target_slot_teacher = teacher_timetables[teacher][day_idx][hour_idx]

                # Eğer yuva zaten boşsa, doğrudan yerleştir ve çık (En kolay senaryo)
                if target_slot_class is None and target_slot_teacher is None:
                    self._place_lesson_safely(teacher, course, class_name, day_idx, hour_idx, class_timetables, teacher_timetables)
                    return True # Başarılı

                # 2. EĞER YUVA DOLUYSA, TAHLİYE OPERASYONUNU DEĞERLENDİR
                lessons_to_evict = []
                if target_slot_class is not None:
                    # Sınıfın o saatte başka dersi var. Bu dersi tahliye etmeliyiz.
                    evicted_lesson = target_slot_class.copy()
                    if evicted_lesson not in lessons_to_evict: lessons_to_evict.append(evicted_lesson)
                
                if target_slot_teacher is not None:
                    # Öğretmenin o saatte başka dersi var. Bu dersi de tahliye etmeliyiz.
                    evicted_lesson = target_slot_teacher.copy()
                    if evicted_lesson not in lessons_to_evict: lessons_to_evict.append(evicted_lesson)
                
                # Kendi dersimizi tahliye etmeye çalışmayalım
                lessons_to_evict = [l for l in lessons_to_evict if l['teacher'] != teacher or l['class'] != class_name]
                if not lessons_to_evict:
                    continue

                # 3. TAHLİYE ET VE YENİ YER BUL
                original_state = (
                    [l.copy() for l in lessons_to_evict],
                    {l['class']: class_timetables[l['class']][day_idx][hour_idx] for l in lessons_to_evict},
                    {l['teacher']: teacher_timetables[l['teacher']][day_idx][hour_idx] for l in lessons_to_evict}
                )
                
                # Önce yuvayı boşalt
                for l in lessons_to_evict:
                    class_timetables[l['class']][day_idx][hour_idx] = None
                    teacher_timetables[l['teacher']][day_idx][hour_idx] = None

                # Tahliye edilen dersler için yeni boş yer ara
                all_evicted_placed = True
                for l_evicted in lessons_to_evict:
                    new_day, new_hour = self._find_empty_slot_for_lesson(
                        l_evicted['teacher'], l_evicted['course'], l_evicted['class'],
                        class_timetables, teacher_timetables
                    )
                    if new_day is not None:
                        self._place_lesson_safely(l_evicted['teacher'], l_evicted['course'], l_evicted['class'], new_day, new_hour, class_timetables, teacher_timetables)
                    else:
                        all_evicted_placed = False
                        break # Bir tanesi bile yerleşemezse operasyon başarısız

                # 4. SONUÇ
                if all_evicted_placed:
                    # Harika! Herkesi taşıdık, şimdi asıl dersimizi boşalan yuvaya koyalım
                    self._place_lesson_safely(teacher, course, class_name, day_idx, hour_idx, class_timetables, teacher_timetables)
                    #print(f"AGRESİF YERLEŞTİRME BAŞARILI: {course} ({class_name}) -> {day_name} {hour_idx+1}. saate yerleşti.")
                    return True # Operasyon başarılı
                else:
                    # Tahliye operasyonu başarısız. Her şeyi eski haline getir.
                    for l_original in original_state[0]:
                        class_timetables[l_original['class']][day_idx][hour_idx] = original_state[1][l_original['class']]
                        teacher_timetables[l_original['teacher']][day_idx][hour_idx] = original_state[2][l_original['teacher']]

        return False # Tüm denemelere rağmen yerleştirilemedi

    # Bu fonksiyonu sınıfınıza ekleyin ve generate_timetable'ın en sonunda çağırın.
    def validate_schedule_consistency(self, class_timetables, teacher_timetables):
        print("\n✅ Tutarlılık kontrolü başlatılıyor...")
        mismatch_count = 0

        for class_name, class_schedule in class_timetables.items():
            for day in range(len(self.days_of_week)):
                for hour in range(self.daily_hours):
                    class_entry = class_schedule[day][hour]
                    # Eğer sınıfın hücresi boşsa, öğretmenin de boş olmalı
                    if not class_entry or not isinstance(class_entry, dict):
                        continue

                    # Sınıf hücresi doluysa, öğretmenin hücresini kontrol et
                    teacher_id = class_entry['teacher']
                    
                    # Öğretmen programında ilgili hücreye bak
                    teacher_entry = teacher_timetables.get(teacher_id, [[None]*self.daily_hours for _ in self.days_of_week])[day][hour]

                    # Karşılaştır
                    if not teacher_entry or not isinstance(teacher_entry, dict):
                        print(f"🔴 UYUMSUZLUK: {class_name} sınıfı {self.days_of_week[day]} {hour+1}. saat -> {class_entry}, ama {teacher_id} tablosunda bu saat BOŞ veya geçersiz!")
                        mismatch_count += 1
                        continue
                    
                    if teacher_entry['class'] != class_name or teacher_entry['course'] != class_entry['course']:
                        print(f"🔴 UYUMSUZLUK: {class_name} sınıfı {self.days_of_week[day]} {hour+1}. saat -> {class_entry}, ama {teacher_id} tablosunda bu saat {teacher_entry}")
                        mismatch_count += 1

        if mismatch_count == 0:
            print("\n✅ Sınıf ve öğretmen tabloları tam uyumlu. Hiçbir tutarsızlık bulunmadı.")
            return True
        else:
            print(f"\n🚨 DİKKAT! Toplam {mismatch_count} adet tutarsızlık tespit edildi. Programı kaydetmeden önce kontrol edin.")
            # İsterseniz burada bir QMessageBox ile kullanıcıyı uyarabilirsiniz.
            QMessageBox.critical(self, "TUTARSIZLIK TESPİT EDİLDİ!",
                                f"Program oluşturma sonrası yapılan kontrolde {mismatch_count} adet uyumsuzluk bulundu.\n"
                                "Bu durum, programda bir hata olduğunu gösterir. Lütfen konsol çıktılarını inceleyin.")
            return False

    # BU YENİ YARDIMCI FONKSİYONU SINIFINIZA EKLEYİN
    def _clear_lesson_safely(self, lesson_info, day_idx, hour_idx, class_timetables, teacher_timetables):
        """
        Verilen bir ders bilgisini hem sınıf hem de öğretmen programından güvenli bir şekilde siler.
        """
        if not isinstance(lesson_info, dict):
            return # Silinecek bir şey yok

        class_name = lesson_info.get("class")
        teacher_id = lesson_info.get("teacher")

        if class_name in class_timetables:
            # Sadece o hücredeki dersin bilgisi aynıysa sil
            if class_timetables[class_name][day_idx][hour_idx] == lesson_info:
                class_timetables[class_name][day_idx][hour_idx] = None

        if teacher_id in teacher_timetables:
            # Sadece o hücredeki dersin bilgisi aynıysa sil
            if teacher_timetables[teacher_id][day_idx][hour_idx] == lesson_info:
                teacher_timetables[teacher_id][day_idx][hour_idx] = None

    def _check_data_readiness(self):
        """
        Program oluşturma öncesi verilerin tutarlılığını kontrol eder.
        YENİ: Öğretmeni atanmamış dersleri tespit eder ve kullanıcıyı uyarır.
        """
        if not self.classes:
            QMessageBox.warning(self, "Veri Eksik", "Henüz hiç sınıf oluşturulmamış.")
            return False
        if not self.teachers:
            QMessageBox.warning(self, "Veri Eksik", "Henüz hiç öğretmen eklenmemiş.")
            return False
        if self.daily_hours == 0:
            QMessageBox.warning(self, "Veri Eksik", "Günlük ders saati belirlenmemiş.")
            return False

        unassigned_courses_warnings = []
        for class_name, class_data in self.classes.items():
            defined_hours = class_data.get('hours', {})
            assigned_teachers = class_data.get('assigned_teachers', {})

            # Saat sayısı tanımlanmış ama öğretmeni atanmamış dersleri bul
            for course, hours in defined_hours.items():
                if hours > 0 and course not in assigned_teachers:
                    unassigned_courses_warnings.append(f"- {class_name} sınıfı, {course} dersi ({hours} saat)")

        # Eğer atanmamış dersler varsa, kullanıcıyı uyar ve onaya göre devam et
        if unassigned_courses_warnings:
            warning_message = "Aşağıdaki derslerin öğretmen ataması yapılmamış.\n"
            warning_message += "Bu dersler programda yer almayacaktır.\n\n"
            warning_message += "\n".join(unassigned_courses_warnings)
            warning_message += "\n\nYine de devam etmek istiyor musunuz?"

            reply = QMessageBox.question(self, 'Eksik Atamalar Tespit Edildi',
                                           warning_message, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

            if reply == QMessageBox.No:
                return False # Kullanıcı devam etmek istemedi

        # Her şey yolundaysa veya kullanıcı devam etmeyi seçtiyse
        return True

    # LÜTFEN MEVCUT _get_lessons_to_place FONKSİYONUNU BU BLOK İLE DEĞİŞTİRİN
    def _get_lessons_to_place(self):
        """Yerleştirilmesi gereken tüm ders saatlerini (ders, öğretmen, sınıf, saat, dağılım) listeler."""
        lessons = []
        for class_name, c_data in self.classes.items():
            hours = c_data.get('hours', {})
            assigned_teachers = c_data.get('assigned_teachers', {})
            distributions = c_data.get('distribution', {}) # YENİ: Dağılım verisini al

            for course, teacher_id in assigned_teachers.items():
                num_hours = hours.get(course, 0)
                distribution = distributions.get(course, []) # YENİ: Derse ait dağılımı al

                if num_hours > 0 and teacher_id in self.teachers:
                    lessons.append({
                        'teacher': teacher_id,
                        'course': course,
                        'class': class_name,
                        'hours': num_hours,
                        'distribution': distribution # YENİ: Dağılımı ekle
                    })
        return lessons
    def generate_html_summary(self, mode='json'):
        from collections import defaultdict

        # Bu metot 'program' modunu desteklemediği için şimdilik kaldırıyoruz.
        # İhtiyaç halinde programdan okuyacak şekilde genişletilebilir.
        # if mode == 'program':
        #     # Programdan veri okuma mantığı buraya eklenmeli
        #     return "<h2>Program Özeti Henüz Uygulanmadı</h2>"

        total_required = 0
        total_assigned = 0
        total_available = 0

        required_by_course = defaultdict(int)
        assigned_by_json = defaultdict(int)
        available_by_course = defaultdict(int)

        lessons = self._get_lessons_to_place()
        for lesson in lessons:
            total_required += lesson['hours']
            required_by_course[lesson['course']] += lesson['hours']

            teacher = self.classes[lesson['class']]["assigned_teachers"].get(lesson['course'])
            if teacher:
                assigned_by_json[lesson['course']] += lesson['hours']
                total_assigned += lesson['hours']

        for teacher, info in self.teachers.items():
            availability = info.get("availability", {})
            # Günlük ders saatinin her gün için aynı olduğunu varsayıyoruz.
            # 5 gün üzerinden hesaplama yapılıyor.
            teacher_available_hours = self.daily_hours * len(self.days_of_week) - sum(len(hours) for hours in availability.values())
            total_available += teacher_available_hours

            for course in info.get("courses", []):
                # Bu hesaplama yanıltıcı olabilir, öğretmenin tüm boş saatlerini o derse atar.
                # Daha doğru bir gösterim için bu kısmı basitleştirebiliriz.
                available_by_course[course] += teacher_available_hours

        html = f"<h2>📊 Planlanan (JSON Verisine Göre)</h2>"
        html += f"<p><b>📌 Toplam ihtiyaç duyulan ders saati:</b> {total_required}</p>"
        html += f"<p><b>✅ Ataması yapılmış (öğretmen belirtilmiş) ders saati:</b> {total_assigned}</p>"
        html += f"<p><b>🧑‍🏫 Öğretmenlerin toplam müsait saati (yaklaşık):</b> {total_available}</p>"
        html += "<p><small>Not: Öğretmenlerin müsait saatleri, derslere bölünmeden toplam olarak gösterilmiştir.</small></p>"

        html += "<table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%;'>"
        html += "<tr style='background-color: #f2f2f2;'><th>Ders</th><th>İhtiyaç (saat)</th><th>Atanmış (saat)</th></tr>"

        all_courses = set(required_by_course) | set(assigned_by_json)
        for course in sorted(all_courses):
            r = required_by_course.get(course, 0)
            a = assigned_by_json.get(course, 0)
            html += f"<tr><td>{course}</td><td style='text-align: center;'>{r}</td><td style='text-align: center;'>{a}</td></tr>"

        html += "</table>"
        return html
    # Bu fonksiyonu mevcut _generate_neighbor_solution ile değiştirin
    # Lütfen bu fonksiyonu kopyalayıp eskisinin üzerine yapıştırın
    def _generate_neighbor_solution(self, class_tables, teacher_tables, unplaced_lessons):
        """
        Komşu çözüm üretmek için güvenli yerleştirme stratejileri kullanır.
        _place_lesson_safely fonksiyonunu esas alır.
        """
        import random
        import copy

        new_class = {k: [row[:] for row in v] for k, v in class_tables.items()}
        new_teacher = {k: [row[:] for row in v] for k, v in teacher_tables.items()}
        new_unplaced = [l.copy() for l in unplaced_lessons]

        if new_unplaced and random.random() < 0.7:
            strategy = "place_unplaced"
        else:
            strategy = random.choice(["move", "swap"])

        if strategy == "place_unplaced" and new_unplaced:
            lesson = random.choice(new_unplaced)
            placed = self._aggressively_try_to_place(lesson, new_class, new_teacher)
            if placed:
                for item in new_unplaced:
                    if item['teacher'] == lesson['teacher'] and item['class'] == lesson['class'] and item['course'] == lesson['course']:
                        item['hours'] -= 1
                        if item['hours'] <= 0:
                            new_unplaced.remove(item)
                        break

        elif strategy == "move":
            class_name = random.choice(list(new_class.keys()))
            day_idx = random.randint(0, len(self.days_of_week)-1)
            hour_idx = random.randint(0, self.daily_hours-1)

            lesson = new_class[class_name][day_idx][hour_idx]
            if isinstance(lesson, dict):
                teacher = lesson["teacher"]
                course = lesson["course"]
                self._clear_lesson_safely(lesson, day_idx, hour_idx, new_class, new_teacher)
                for d in range(len(self.days_of_week)):
                    for h in range(self.daily_hours):
                        if new_class[class_name][d][h] is None and new_teacher[teacher][d][h] is None:
                            self._place_lesson_safely(teacher, course, class_name, d, h, new_class, new_teacher)
                            return new_class, new_teacher, new_unplaced
                self._place_lesson_safely(teacher, course, class_name, day_idx, hour_idx, new_class, new_teacher)

        elif strategy == "swap":
            class_name = random.choice(list(new_class.keys()))
            schedule = new_class[class_name]
            filled = [(d, h) for d in range(len(self.days_of_week)) for h in range(self.daily_hours) if isinstance(schedule[d][h], dict)]

            if len(filled) >= 2:
                (d1, h1), (d2, h2) = random.sample(filled, 2)
                l1, l2 = schedule[d1][h1], schedule[d2][h2]
                t1, t2 = l1["teacher"], l2["teacher"]

                if new_teacher[t1][d2][h2] is None and new_teacher[t2][d1][h1] is None:
                    self._clear_lesson_safely(l1, d1, h1, new_class, new_teacher)
                    self._clear_lesson_safely(l2, d2, h2, new_class, new_teacher)
                    self._place_lesson_safely(t1, l1["course"], class_name, d2, h2, new_class, new_teacher)
                    self._place_lesson_safely(t2, l2["course"], class_name, d1, h1, new_class, new_teacher)

        return new_class, new_teacher, new_unplaced

    # ✅ _generate_neighbor_solution fonksiyonu da güvenli hale getirildi.

    def generate_program_summary_html(self):
        """Oluşturulan ders programını analiz eder ve HTML formatında özet döndürür."""
        from collections import defaultdict

        # 1. Yerleştirilen dersleri analiz et
        placed_by_course = defaultdict(int)
        total_placed_hours = 0
        if hasattr(self, 'class_timetables'):
            for class_name, timetable in self.class_timetables.items():
                for day_list in timetable:
                    for lesson_info in day_list:
                        if isinstance(lesson_info, dict):
                            total_placed_hours += 1
                            course = lesson_info.get('course')
                            if course:
                                placed_by_course[course] += 1

        # 2. İhtiyaç duyulan dersleri hesapla (karşılaştırma için)
        required_by_course = defaultdict(int)
        total_required_hours = 0
        all_required_lessons = self._get_lessons_to_place()
        for lesson in all_required_lessons:
            total_required_hours += lesson['hours']
            required_by_course[lesson['course']] += lesson['hours']

        # 3. HTML çıktısını oluştur
        html = f"<h2>📊 Oluşturulan Program Analizi</h2>"
        placement_rate = (total_placed_hours / total_required_hours * 100) if total_required_hours > 0 else 0
        html += f"<p><b>📌 Toplam İstenen Saat:</b> {total_required_hours}</p>"
        html += f"<p><b>✅ Programa Yerleştirilen Saat:</b> {total_placed_hours}</p>"
        html += f"<p><b>📈 Yerleştirme Oranı:</b> {placement_rate:.2f}%</p>"
        html += "<hr>"

        # Ders Bazlı Karşılaştırma Tablosu
        html += "<h3>Ders Bazlı Karşılaştırma</h3>"
        html += "<table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%;'>"
        html += "<tr style='background-color: #f2f2f2;'><th>Ders</th><th>İstenen Saat</th><th>Yerleştirilen Saat</th><th>Fark</th></tr>"

        all_courses = sorted(list(set(required_by_course.keys()) | set(placed_by_course.keys())))
        for course in all_courses:
            required = required_by_course.get(course, 0)
            placed = placed_by_course.get(course, 0)
            diff = placed - required
            color = "green" if diff == 0 else "red" if diff < 0 else "blue"
            html += f"<tr><td>{course}</td><td style='text-align: center;'>{required}</td><td style='text-align: center;'>{placed}</td><td style='text-align: center; color: {color}; font-weight: bold;'>{diff}</td></tr>"
        html += "</table>"
        html += "<br>"

        # Yerleştirilemeyen Dersler Tablosu
        unplaced_lessons = self.last_run_unplaced_lessons
        if unplaced_lessons:
            html += f"<h3 style='color: red;'>Yerleştirilemeyen Dersler ({len(unplaced_lessons)} adet)</h3>"
            html += "<table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%;'>"
            html += "<tr style='background-color: #f2f2f2;'><th>Sınıf</th><th>Ders</th><th>Öğretmen</th><th>Eksik Saat</th></tr>"
            for lesson in unplaced_lessons:
                html += f"<tr><td>{lesson['class']}</td><td>{lesson['course']}</td><td>{lesson['teacher']}</td><td style='text-align: center;'>{lesson['hours']}</td></tr>"
            html += "</table>"

        return html

    def show_total_lesson_summary(self):
        """JSON verisi ve oluşturulan programa göre saat özetlerini sekmeli olarak gösterir."""
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QTextEdit, QPushButton, QTabWidget, QWidget

        dialog = QDialog(self)
        dialog.setWindowTitle("Ders Saati Özeti (Planlanan vs. Oluşturulan)")
        dialog.setMinimumSize(750, 500)
        layout = QVBoxLayout(dialog)

        tabs = QTabWidget()
        layout.addWidget(tabs)

        # Sekme 1: Planlanan (JSON'a göre)
        json_tab = QWidget()
        json_layout = QVBoxLayout(json_tab)
        json_text = QTextEdit()
        json_text.setReadOnly(True)
        json_text.setHtml(self.generate_html_summary('json')) # Bu eski fonksiyonu kullanıyor
        json_layout.addWidget(json_text)
        tabs.addTab(json_tab, "Planlanan (Giriş Verileri)")

        # Sekme 2: Oluşturulan Program Analizi
        program_tab = QWidget()
        program_layout = QVBoxLayout(program_tab)
        program_text = QTextEdit()
        program_text.setReadOnly(True)

        # Programın daha önce oluşturulup oluşturulmadığını kontrol et
        if hasattr(self, 'class_timetables') and self.class_timetables:
            # Eğer oluşturulduysa, yeni yazdığımız analiz fonksiyonunu çağır
            program_text.setHtml(self.generate_program_summary_html())
        else:
            # Eğer henüz oluşturulmadıysa, kullanıcıya bilgi ver
            program_text.setHtml("<h2>Oluşturulan Program Analizi</h2><p>Bu özeti görmek için önce 'Ders Programı Oluştur' butonunu kullanarak bir program oluşturmalısınız.</p>")
        
        program_layout.addWidget(program_text)
        tabs.addTab(program_tab, "Oluşturulan Program Analizi")

        # Kapat Butonu
        close_button = QPushButton("Kapat")
        close_button.clicked.connect(dialog.accept)
        layout.addWidget(close_button)

        dialog.exec_()

    def _calculate_teacher_total_hours(self):
        """Her öğretmenin haftalık toplam atanmış ders saatini hesaplar."""
        teacher_hours = {t_id: 0 for t_id in self.teachers}
        for t_id, t_data in self.teachers.items():
            if 'assignments' in t_data:
                for course, classes in t_data['assignments'].items():
                    for class_name in classes:
                        if class_name in self.classes and isinstance(self.classes[class_name], dict):
                            teacher_hours[t_id] += self.classes[class_name].get('hours', {}).get(course, 0)
        return teacher_hours

    def _find_and_place_lesson(self, teacher_id, course, class_name, hours_to_place,
                            class_timetables, teacher_timetables):
        """Bir dersi uygun bir zaman dilimine yerleştirir. Gün sırası rastgeledir."""
        teacher_availability = self.teachers[teacher_id].get('availability', {})

        shuffled_days = list(self.days_of_week)
        random.shuffle(shuffled_days)

        for day_name in shuffled_days:
            day_idx = self.days_of_week.index(day_name)
            unavailable_hours_for_day = teacher_availability.get(day_name, [])

            possible_start_hours = list(range(self.daily_hours - hours_to_place + 1))
            random.shuffle(possible_start_hours)

            for start_hour in possible_start_hours:
                can_place = True
                for i in range(hours_to_place):
                    hour_idx = start_hour + i
                    hour_human = hour_idx + 1

                    if hour_human in unavailable_hours_for_day:
                        can_place = False
                        break
                    if teacher_timetables[teacher_id][day_idx][hour_idx] is not None:
                        can_place = False
                        break
                    if class_timetables[class_name][day_idx][hour_idx] is not None:
                        can_place = False
                        break

                if can_place:
                    for i in range(hours_to_place):
                        hour_idx = start_hour + i
                        self._place_lesson_safely(
                            teacher=teacher_id,
                            course=course,
                            class_name=class_name,
                            day=day_idx,
                            hour=hour_idx,
                            class_timetables=class_timetables,
                            teacher_timetables=teacher_timetables
                        )
                    return True

        return False
    
    def _aggressively_try_to_place(self, lesson, class_tables, teacher_tables):
        """
        Yerleştirilemeyen dersleri daha fazla iterasyonla zorlama yerleştirme denemesi.
        Artık _place_lesson_safely fonksiyonunu kullanarak yerleştirir.
        """
        teacher = lesson['teacher']
        course = lesson['course']
        class_name = lesson['class']
        hours = lesson['hours']

        teacher_availability = self.teachers[teacher].get('availability', {})
        days = list(self.days_of_week)
        random.shuffle(days)

        for day_name in days:
            day_idx = self.days_of_week.index(day_name)
            unavailable = teacher_availability.get(day_name, [])
            for hour in range(self.daily_hours):
                if (hour + 1) in unavailable:
                    continue
                if teacher_tables[teacher][day_idx][hour] is None and class_tables[class_name][day_idx][hour] is None:
                    self._place_lesson_safely(
                        teacher=teacher,
                        course=course,
                        class_name=class_name,
                        day=day_idx,
                        hour=hour,
                        class_timetables=class_tables,
                        teacher_timetables=teacher_tables
                    )
                    hours -= 1
                    if hours <= 0:
                        return True
        return False


    def validate_schedule_consistency(self, class_timetables, teacher_timetables):
        print("\n✅ Tutarlılık kontrolü başlatılıyor...")
        mismatch_count = 0

        for class_name, class_schedule in class_timetables.items():
            for day in range(len(self.days_of_week)):
                for hour in range(self.daily_hours):
                    class_entry = class_schedule[day][hour]
                    if not class_entry or not isinstance(class_entry, dict):
                        continue

                    teacher_id = class_entry['teacher']
                    teacher_entry = teacher_timetables.get(teacher_id, [[None]*self.daily_hours for _ in self.days_of_week])[day][hour]

                    if not teacher_entry or not isinstance(teacher_entry, dict):
                        print(f"UYUMSUZLUK: {class_name} sınıfı {self.days_of_week[day]} {hour+1}. saat → {class_entry}, ama {teacher_id} tablosunda bu saat boş veya geçersiz")
                        mismatch_count += 1
                        continue

                    if teacher_entry['class'] != class_name or teacher_entry['course'] != class_entry['course']:
                        print(f"UYUMSUZLUK: {class_name} sınıfı {self.days_of_week[day]} {hour+1}. saat → {class_entry}, ama {teacher_id} tablosunda bu saat {teacher_entry}")
                        mismatch_count += 1

        if mismatch_count == 0:
            print("\n✅ Sınıf ve öğretmen tabloları tam uyumlu.")
        else:
            print(f"\n🔴 TERS YÖNLÜ UYUŞMA SORUNLARI TESPİT EDİLDİ: {mismatch_count} adet.")

    def _sanitize_timetable(self, timetable):
        for key in timetable:
            for day in range(len(self.days_of_week)):
                for hour in range(self.daily_hours):
                    val = timetable[key][day][hour]
                    if not isinstance(val, dict):
                        timetable[key][day][hour] = None


    def _report_and_save_timetable(self, class_timetables, teacher_timetables, unplaced_lessons, stats):
        """Oluşturulan programı kullanıcıya bildirir ve Excel'e kaydeder."""

        # Sonuç Mesajı Oluşturma
        total_lessons = stats['total']
        placed_lessons = stats['placed']
        unplaced_count = len(unplaced_lessons)
        success_rate = (placed_lessons / total_lessons * 100) if total_lessons > 0 else 0

        result_message = f"Ders Programı Oluşturma Tamamlandı!\n\n"
        result_message += f"Toplam Ders Saati: {total_lessons}\n"
        result_message += f"Yerleştirilen Ders Saati: {placed_lessons}\n"
        result_message += f"Başarı Oranı: {success_rate:.2f}%\n\n"

        if unplaced_lessons:
            result_message += f"UYARI: Aşağıdaki {unplaced_count} ders/saat yerleştirilemedi:\n"
            for item in unplaced_lessons[:10]: # İlk 10 tanesini göster
                 result_message += f"- {item['class']} Sınıfı, {item['course']} Dersi ({item['teacher']}), {item['hours']} saat\n"
            if unplaced_count > 10: result_message += "- ... (ve diğerleri)\n"
            result_message += "\nOluşturulan programda boşluklar olacaktır.\n"
        else:
            result_message += "Tüm dersler başarıyla yerleştirildi!\n"

        result_message += "\nProgramı Excel dosyası olarak kaydetmek ister misiniz?"

        reply = QMessageBox.information(self, "Program Oluşturma Sonucu", result_message,
                                        QMessageBox.Save | QMessageBox.Cancel, QMessageBox.Save)

        if reply == QMessageBox.Save:
            self._save_timetables_to_excel(class_timetables, teacher_timetables)
   # def format_lesson_cell(info, for_teacher=True):
     #   if isinstance(info, dict):
      #      course = info.get("course", "")
       #     other = info.get("class" if for_teacher else "teacher", "")
       #     return f"{course} ({other})"
       # return info if info else "-"


    def _save_timetables_to_excel(self, class_timetables, teacher_timetables):
        """Zaman çizelgelerini Excel dosyasına kaydeder."""
        if not OPENPYXL_AVAILABLE: return # Kütüphane yoksa çık

        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(self, "Ders Programını Kaydet", "",
                                                   "Excel Dosyaları (*.xlsx);;All Files (*)", options=options)
        if not file_path: return
        if not file_path.lower().endswith('.xlsx'): file_path += '.xlsx'

        try:
            wb = openpyxl.Workbook()
            # Varsayılan 'Sheet' sayfasını kaldır
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            header_font = Font(bold=True, name='Calibri', size=11)
            cell_font = Font(name='Calibri', size=10)
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border_side = Side(border_style="thin", color="000000")
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

            # === Sınıf Programları Sayfası ===
            ws_classes = wb.create_sheet("Sınıf Programları")
            row_offset = 1 # Başlangıç satırı

            sorted_classes = sorted(class_timetables.keys())
            num_classes = len(sorted_classes)
            tables_per_row = 2 # A4'e sığdırmak için yan yana 2 tablo hedefi
            num_rows_per_table = self.daily_hours + 2 # Başlık + Saatler + Sınıf Adı
            col_width = 18 # Sütun genişliği (ayarlanabilir)

            for i, class_name in enumerate(sorted_classes):
                timetable = class_timetables[class_name]

                # Tablonun başlangıç sütununu hesapla (yan yana 2 tablo için)
                table_col_start = 1 + (i % tables_per_row) * (len(self.days_of_week) + 1) # +1 saat sütunu için
                # Tablonun başlangıç satırını hesapla
                table_row_start = row_offset + math.floor(i / tables_per_row) * (num_rows_per_table + 1) # +1 boşluk için

                # Sınıf Adı Başlığı
                class_header_cell = ws_classes.cell(row=table_row_start, column=table_col_start,
                                                   value=f"{class_name} Sınıfı Ders Programı")
                class_header_cell.font = Font(bold=True, size=12)
                class_header_cell.alignment = center_alignment
                ws_classes.merge_cells(start_row=table_row_start, start_column=table_col_start,
                                       end_row=table_row_start, end_column=table_col_start + len(self.days_of_week))

                # Gün Başlıkları
                header_row = table_row_start + 1
                ws_classes.cell(row=header_row, column=table_col_start).value = "Saat" # İlk sütun başlığı
                ws_classes.cell(row=header_row, column=table_col_start).font = header_font
                ws_classes.cell(row=header_row, column=table_col_start).alignment = center_alignment
                ws_classes.cell(row=header_row, column=table_col_start).border = thin_border
                ws_classes.column_dimensions[get_column_letter(table_col_start)].width = 8 # Saat sütunu dar

                for c_idx, day in enumerate(self.days_of_week):
                    col = table_col_start + 1 + c_idx
                    cell = ws_classes.cell(row=header_row, column=col, value=day)
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    ws_classes.column_dimensions[get_column_letter(col)].width = col_width

                # Zaman Çizelgesi Verileri
                for r_idx in range(self.daily_hours):
                    row = table_row_start + 2 + r_idx
                    hour_cell = ws_classes.cell(row=row, column=table_col_start, value=f"{r_idx + 1}. Ders")
                    hour_cell.font = header_font
                    hour_cell.alignment = center_alignment
                    hour_cell.border = thin_border
                    ws_classes.row_dimensions[row].height = 35 # Satır yüksekliği (ayarlanabilir)

                    for c_idx in range(len(self.days_of_week)):
                        col = table_col_start + 1 + c_idx
                        cell = ws_classes.cell(row=row, column=col)
                        lesson_info = timetable[c_idx][r_idx]
                        if isinstance(lesson_info, dict):
                            course = lesson_info.get("course", "")
                            teacher = lesson_info.get("teacher", "")
                            cell.value = f"{course} ({teacher})"
                        else:
                            cell.value = lesson_info if lesson_info else "-"
                        cell.font = cell_font
                        cell.alignment = center_alignment
                        cell.border = thin_border


            # Sayfa Ayarları (Sınıf) - A4'e sığdırma denemesi
            ws_classes.page_setup.orientation = ws_classes.ORIENTATION_LANDSCAPE
            ws_classes.page_setup.paperSize = ws_classes.PAPERSIZE_A4
            # Fit to page genişlikte 1 sayfa, yükseklikte otomatik
            ws_classes.page_setup.fitToWidth = 1
            ws_classes.page_setup.fitToHeight = 0 # Otomatik yükseklik
            ws_classes.print_options.horizontalCentered = True
            #ws_classes.print_options.verticalCentered = True # Opsiyonel


            # === Öğretmen Programları Sayfası ===
            ws_teachers = wb.create_sheet("Öğretmen Programları")
            # Öğretmen programları için benzer bir döngü ve formatlama yapılabilir.
            # Şimdilik basit bir liste halinde yazalım:
            row_offset = 1
            sorted_teachers = sorted(teacher_timetables.keys())
            for i, teacher_id in enumerate(sorted_teachers):
                 timetable = teacher_timetables[teacher_id]
                 # Tablonun başlangıç sütununu hesapla
                 table_col_start = 1 + (i % tables_per_row) * (len(self.days_of_week) + 1)
                 # Tablonun başlangıç satırını hesapla
                 table_row_start = row_offset + math.floor(i / tables_per_row) * (num_rows_per_table + 1)

                 # Öğretmen Adı Başlığı
                 teacher_header_cell = ws_teachers.cell(row=table_row_start, column=table_col_start, value=f"{teacher_id} Ders Programı")
                 teacher_header_cell.font = Font(bold=True, size=12)
                 teacher_header_cell.alignment = center_alignment
                 ws_teachers.merge_cells(start_row=table_row_start, start_column=table_col_start, end_row=table_row_start, end_column=table_col_start + len(self.days_of_week))

                 # Gün Başlıkları
                 header_row = table_row_start + 1
                 ws_teachers.cell(row=header_row, column=table_col_start).value = "Saat"
                 ws_teachers.cell(row=header_row, column=table_col_start).font = header_font
                 ws_teachers.cell(row=header_row, column=table_col_start).alignment = center_alignment
                 ws_teachers.cell(row=header_row, column=table_col_start).border = thin_border
                 ws_teachers.column_dimensions[get_column_letter(table_col_start)].width = 8

                 for c_idx, day in enumerate(self.days_of_week):
                     col = table_col_start + 1 + c_idx
                     cell = ws_teachers.cell(row=header_row, column=col, value=day)
                     cell.font = header_font
                     cell.alignment = center_alignment
                     cell.border = thin_border
                     ws_teachers.column_dimensions[get_column_letter(col)].width = col_width

                 # Zaman Çizelgesi Verileri
                 for r_idx in range(self.daily_hours):
                     row = table_row_start + 2 + r_idx
                     hour_cell = ws_teachers.cell(row=row, column=table_col_start, value=f"{r_idx + 1}. Ders")
                     hour_cell.font = header_font
                     hour_cell.alignment = center_alignment
                     hour_cell.border = thin_border
                     ws_teachers.row_dimensions[row].height = 35

                     for c_idx in range(len(self.days_of_week)):
                         col = table_col_start + 1 + c_idx
                         cell = ws_teachers.cell(row=row, column=col)
                         lesson_info = timetable[c_idx][r_idx]
                         if isinstance(lesson_info, dict):
                            course = lesson_info.get("course", "")
                            class_name = lesson_info.get("class", "")
                            cell.value = f"{course} ({class_name})"
                         else:
                            cell.value = lesson_info if lesson_info else "-"
                         cell.font = cell_font
                         cell.alignment = center_alignment
                         cell.border = thin_border

            # Sayfa Ayarları (Öğretmen)
            ws_teachers.page_setup.orientation = ws_teachers.ORIENTATION_LANDSCAPE
            ws_teachers.page_setup.paperSize = ws_teachers.PAPERSIZE_A4
            ws_teachers.page_setup.fitToWidth = 1
            ws_teachers.page_setup.fitToHeight = 0
            ws_teachers.print_options.horizontalCentered = True

            # Workbook'u Kaydet
            wb.save(file_path)
            QMessageBox.information(self, "Başarılı", f"Ders programı başarıyla '{os.path.basename(file_path)}' dosyasına kaydedildi.")

        except PermissionError:
             QMessageBox.critical(self, "Hata", f"Dosya yazma izni hatası.\n'{file_path}' dosyası başka bir program tarafından kullanılıyor olabilir veya yazma izniniz olmayabilir.")
        except Exception as e:
            QMessageBox.critical(self, "Excel Kaydetme Hatası", f"Program Excel'e kaydedilirken bir hata oluştu: {str(e)}")
    


    # <<< ============================================= >>>
    # <<<      DERS PROGRAMI OLUŞTURMA BÖLÜMÜ SONU     >>>
    # <<< ============================================= >>>


if __name__ == '__main__':
    import sys
    app = QApplication(sys.argv)
    scheduler = SchoolSchedulerApp()
    scheduler.setWindowTitle('Okul Ders Programı Hazırlık Aracı By Serkan Bıçakcı')
    # Pencere boyutunu biraz daha genişletelim
    scheduler.setGeometry(100, 100, 800, 700) # Genişlik artırıldı
    scheduler.show()
    sys.exit(app.exec_())